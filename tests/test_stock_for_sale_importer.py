from datetime import date
from pathlib import Path

from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from app.db.models import Base, StockForSaleFact, StockForSaleSource
from app.importers.stock_for_sale import import_stock_for_sale, parse_snapshot_month_from_filename, parse_stock_for_sale_file


def test_parse_snapshot_month_from_filename() -> None:
    period = parse_snapshot_month_from_filename(
        Path("\u041e\u0441\u0442\u0430\u0442\u043a\u0438 \u0432 \u043f\u0440\u043e\u0434\u0430\u0436\u0435_30.04.2026.xlsx"),
    )

    assert period == date(2026, 4, 1)


def test_parse_stock_for_sale_file_reads_totals_and_floor_rows(tmp_path) -> None:
    file_path = tmp_path / "\u041e\u0441\u0442\u0430\u0442\u043a\u0438 \u0432 \u043f\u0440\u043e\u0434\u0430\u0436\u0435_30.04.2026.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "\u041e\u0441\u0442\u0430\u0442\u043a\u0438 \u0432 \u043f\u0440\u043e\u0434\u0430\u0436\u0435"
    sheet.cell(row=4, column=2, value=date(2026, 4, 30))
    sheet.cell(row=5, column=2, value="\u0432 \u043f\u0440\u043e\u0434\u0430\u0436\u0435:")
    sheet.cell(row=5, column=3, value="\u0421\u0443\u043c\u043c\u0430 \u0414\u0414\u0423")
    sheet.cell(row=5, column=4, value="\u041d\u0430\u0446\u0435\u043d\u043a\u0430 \u0414\u0423\u041f\u0422")
    sheet.cell(row=5, column=5, value="\u043a\u0432.\u043c")
    sheet.cell(row=5, column=6, value="\u0448\u0442.")
    sheet.cell(row=5, column=7, value="\u0426\u0415\u041d\u0410 \u0437\u0430 \u043c2 \u0414\u0414\u0423")
    sheet.cell(row=5, column=8, value="\u0426\u0415\u041d\u0410 \u0437\u0430 \u043c2 \u0414\u0423\u041f\u0422")
    sheet.cell(row=6, column=2, value="\u0412\u0441\u0435\u0433\u043e")
    sheet.cell(row=6, column=3, value=1000)
    sheet.cell(row=6, column=4, value=400)
    sheet.cell(row=6, column=5, value=10)
    sheet.cell(row=6, column=6, value=2)
    sheet.cell(row=7, column=3, value=1400)
    sheet.cell(row=11, column=2, value="1 \u044d\u0442\u0430\u0436 \u0432 \u0440\u0430\u0431\u043e\u0442\u0435")
    sheet.cell(row=11, column=3, value=300)
    sheet.cell(row=11, column=4, value=120)
    sheet.cell(row=11, column=5, value=3)
    sheet.cell(row=11, column=6, value=1)
    sheet.cell(row=11, column=7, value=100)
    sheet.cell(row=11, column=8, value=40)
    sheet.cell(row=12, column=2, value="\u0421\u0417 \u043e\u0431\u0432, 7 \u044d\u0442\u0430\u0436")
    sheet.cell(row=12, column=3, value=700)
    sheet.cell(row=12, column=5, value=7)
    sheet.cell(row=12, column=6, value=1)
    workbook.save(file_path)

    source, facts = parse_stock_for_sale_file(file_path, "obvodny")

    assert source.snapshot_month == date(2026, 4, 1)
    assert source.snapshot_date == date(2026, 4, 30)
    assert len(facts) == 4
    assert facts[0].row_type == "total"
    assert facts[0].total_amount == 1400
    assert facts[1].row_type == "total_with_markup"
    assert facts[1].row_label == "\u0412\u0441\u0435\u0433\u043e \u0441 \u043d\u0430\u0446\u0435\u043d\u043a\u043e\u0439 \u0414\u0423\u041f\u0422"
    assert facts[2].property_type == "first_floor"
    assert facts[2].floor_number == 1
    assert facts[2].is_in_work is True
    assert facts[2].total_price_per_sqm == 140
    assert facts[3].property_type == "developer_balance"
    assert facts[3].floor_number == 7


def test_import_stock_for_sale_is_idempotent(tmp_path) -> None:
    file_path = tmp_path / "\u041e\u0441\u0442\u0430\u0442\u043a\u0438 \u0432 \u043f\u0440\u043e\u0434\u0430\u0436\u0435_31.03.2026.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.cell(row=4, column=2, value=date(2026, 3, 31))
    sheet.cell(row=6, column=2, value="\u0430\u043f\u0430\u0440\u0442\u0430\u043c\u0435\u043d\u0442\u044b")
    sheet.cell(row=6, column=3, value=500)
    sheet.cell(row=6, column=4, value=100)
    sheet.cell(row=6, column=5, value=5)
    sheet.cell(row=6, column=6, value=2)
    workbook.save(file_path)

    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)

    with Session(engine) as session:
        first = import_stock_for_sale(session, tmp_path, "obvodny")
        second = import_stock_for_sale(session, tmp_path, "obvodny")

    with Session(engine) as session:
        assert first.files == 1
        assert first.sources == 1
        assert first.facts == 1
        assert second.facts == 1
        assert session.query(StockForSaleSource).count() == 1
        assert session.query(StockForSaleFact).count() == 1
        fact = session.query(StockForSaleFact).one()
        assert fact.property_type == "apartment"
        assert fact.total_amount == 600
