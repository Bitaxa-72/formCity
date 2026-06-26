from datetime import date
from pathlib import Path

from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from app.db.models import (
    Base,
    DebtBookingDeviation,
    DebtBookingItem,
    DebtBookingMonthlyValue,
    DebtBookingRefusal,
    DebtBookingSource,
)
from app.importers.debt_bookings import import_debt_bookings, parse_debt_bookings_file, parse_snapshot_date_from_filename


def build_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Отчет о ДЗ и Бронях"
    sheet.cell(row=3, column=6, value="Оплаты по 21.05.2026 учтены")
    sheet.cell(row=3, column=8, value="Особый клиент")
    sheet.cell(row=3, column=9, value="№ помещения")
    sheet.cell(row=3, column=10, value="Итого")
    sheet.cell(row=3, column=11, value=date(2026, 1, 1))
    sheet.cell(row=3, column=12, value=date(2026, 2, 1))
    sheet.cell(row=3, column=23, value="Комментарии")
    sheet.cell(row=3, column=24, value="Контакты")

    sheet.cell(row=4, column=6, value="ИТОГО")
    sheet.cell(row=4, column=10, value=1000)
    sheet.cell(row=4, column=11, value=100)
    sheet.cell(row=5, column=6, value="Зарегистрировано")
    sheet.cell(row=5, column=10, value=500)
    sheet.cell(row=6, column=6, value="Иванов Иван")
    sheet.cell(row=6, column=7, value="Менеджер")
    sheet.cell(row=6, column=8, value="+")
    sheet.cell(row=6, column=9, value="4.12")
    sheet.cell(row=6, column=10, value=300)
    sheet.cell(row=6, column=11, value=150)
    sheet.cell(row=6, column=23, value="Комментарий")
    sheet.cell(row=6, column=24, value="+79990000000")

    deviations = workbook.create_sheet("Отклонения")
    deviations.cell(row=2, column=4, value="План на май 2026")
    deviations.cell(row=2, column=5, value="План на май 2026 скорректированный")
    deviations.cell(row=2, column=6, value="Комментарии")
    deviations.cell(row=2, column=8, value="Факт оплат")
    deviations.cell(row=2, column=9, value="Остаток")
    deviations.cell(row=2, column=10, value="Комментарии")
    deviations.cell(row=3, column=2, value="ИТОГО")
    deviations.cell(row=3, column=4, value=1000)
    deviations.cell(row=3, column=8, value=800)
    deviations.cell(row=4, column=2, value="Зарегистрировано")
    deviations.cell(row=5, column=2, value="Иванов Иван")
    deviations.cell(row=5, column=3, value="4.12")
    deviations.cell(row=5, column=4, value=300)
    deviations.cell(row=5, column=5, value=350)
    deviations.cell(row=5, column=6, value="Плановый комментарий")
    deviations.cell(row=5, column=8, value=200)
    deviations.cell(row=5, column=9, value=-150)
    deviations.cell(row=5, column=10, value="Фактический комментарий")

    refusals = workbook.create_sheet("Отказы")
    refusals.append(["ФИО", "Статус", "Площадь", "Номер", "Стоимость 100%", "Форма оплаты", "Причина", "АН", "Менеджер"])
    refusals.append(["Петров Петр", "Отказ", 42.5, "7.10", 12000000, "100%", "Причина", "АН", "Менеджер"])

    workbook.save(path)


def test_parse_snapshot_date_from_filename() -> None:
    assert parse_snapshot_date_from_filename(Path("Отчет о ДЗ и Бронях_апрель 2026.xlsx")) == date(2026, 4, 30)
    assert parse_snapshot_date_from_filename(Path("Отчет о ДЗ и Бронях_по 21.05.2026.xlsx")) == date(2026, 5, 21)


def test_parse_debt_bookings_file_reads_all_sheets(tmp_path) -> None:
    file_path = tmp_path / "Отчет о ДЗ и Бронях.xlsx"
    build_workbook(file_path)

    source, items, monthly_values, deviations, refusals = parse_debt_bookings_file(file_path, "obvodny")

    assert source.snapshot_date == date(2026, 5, 21)
    assert source.snapshot_month == date(2026, 5, 1)
    assert len(items) == 3
    assert len(monthly_values) == 2
    assert len(deviations) == 3
    assert len(refusals) == 1

    item = next(value for value in items if value.row_type == "detail")
    assert item.item_kind == "registered"
    assert item.client_name == "Иванов Иван"
    assert item.manager_name == "Менеджер"
    assert item.is_special_client is True
    assert item.unit_number == "4.12"
    assert item.total_amount == 300
    assert item.is_sensitive is True
    assert item.sensitive_fields == {
        "client_name": True,
        "manager_name": True,
        "unit_number": True,
        "comments": True,
        "contacts": True,
    }

    assert {(value.item_kind, value.row_type, value.period_month, value.value) for value in monthly_values} == {
        ("total", "category", date(2026, 1, 1), 100),
        ("registered", "detail", date(2026, 1, 1), 150),
    }

    deviation = next(value for value in deviations if value.row_type == "detail")
    assert deviation.period_month == date(2026, 5, 1)
    assert deviation.client_name == "Иванов Иван"
    assert deviation.plan_amount == 300
    assert deviation.updated_plan_amount == 350
    assert deviation.fact_payment_amount == 200
    assert deviation.remaining_amount == -150
    assert deviation.is_sensitive is True

    refusal = refusals[0]
    assert refusal.customer_name == "Петров Петр"
    assert refusal.status == "Отказ"
    assert refusal.area_sqm == 42.5
    assert refusal.full_price_amount == 12000000
    assert refusal.is_sensitive is True


def test_import_debt_bookings_is_idempotent(tmp_path) -> None:
    file_path = tmp_path / "Отчет о ДЗ и Бронях.xlsx"
    build_workbook(file_path)

    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)

    with Session(engine) as session:
        first = import_debt_bookings(session, tmp_path, "obvodny")
        second = import_debt_bookings(session, tmp_path, "obvodny")

    with Session(engine) as session:
        assert first.files == 1
        assert first.sources == 1
        assert first.items == 3
        assert second.items == 3
        assert session.query(DebtBookingSource).count() == 1
        assert session.query(DebtBookingItem).count() == 3
        assert session.query(DebtBookingMonthlyValue).count() == 2
        assert session.query(DebtBookingDeviation).count() == 3
        assert session.query(DebtBookingRefusal).count() == 1
