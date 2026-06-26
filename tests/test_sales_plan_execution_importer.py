from datetime import date
from pathlib import Path

from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from app.db.models import Base, SalesPlanExecutionFact, SalesPlanExecutionSource
from app.importers.sales_plan_execution import (
    import_sales_plan_execution,
    parse_sales_plan_execution_file,
    parse_snapshot_month_from_filename,
)


def test_parse_snapshot_month_from_filename() -> None:
    period = parse_snapshot_month_from_filename(
        Path("Отчет об исполнении плана продаж (Обв.118)_30.04.2026.xlsx"),
    )

    assert period == date(2026, 4, 1)


def test_parse_sales_plan_execution_file_reads_all_blocks(tmp_path) -> None:
    file_path = tmp_path / "Отчет об исполнении плана продаж (Обв.118)_30.04.2026.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Обводный 118"
    sheet.cell(row=2, column=4, value="Отчет об исполнении плана продаж проекта Обводный 118 на 30.04.2026")

    sheet.cell(row=4, column=3, value="ИТОГО ПРОЕКТ")
    sheet.cell(row=4, column=4, value="ПЛАН")
    sheet.cell(row=4, column=5, value="ФАКТ")
    sheet.cell(row=4, column=6, value="Отклонение")
    sheet.cell(row=4, column=8, value="ФАКТ+ПРОГНОЗ")
    sheet.cell(row=4, column=9, value="Отклонение")
    sheet.cell(row=5, column=3, value="Продажи, руб.")
    sheet.cell(row=5, column=4, value=1000)
    sheet.cell(row=5, column=5, value=800)
    sheet.cell(row=5, column=6, value=-200)
    sheet.cell(row=5, column=8, value=1200)
    sheet.cell(row=5, column=9, value=200)
    sheet.cell(row=6, column=3, value="в т.ч. на Застройщика, руб.")
    sheet.cell(row=6, column=4, value=700)
    sheet.cell(row=6, column=5, value=600)
    sheet.cell(row=8, column=3, value="объем законтрактованных площадей, м2")
    sheet.cell(row=8, column=4, value=10)
    sheet.cell(row=8, column=5, value=8)

    sheet.cell(row=16, column=3, value="АПАРТАМЕНТЫ")
    sheet.cell(row=16, column=4, value="ПЛАН")
    sheet.cell(row=16, column=5, value="ФАКТ")
    sheet.cell(row=16, column=6, value="Отклонение")
    sheet.cell(row=16, column=8, value="ФАКТ+ПРОГНОЗ")
    sheet.cell(row=16, column=9, value="Отклонение")
    sheet.cell(row=17, column=3, value="Продажи, руб.")
    sheet.cell(row=17, column=4, value=900)
    sheet.cell(row=17, column=5, value=750)

    sheet.cell(row=39, column=3, value=date(2026, 4, 1))
    sheet.cell(row=39, column=4, value="ПЛАН")
    sheet.cell(row=39, column=5, value="ПРОГНОЗ")
    sheet.cell(row=39, column=6, value="ФАКТ")
    sheet.cell(row=39, column=7, value="РАЗНИЦА\nФАКТ - ПРОГНОЗ")
    sheet.cell(row=40, column=3, value="Продажи, руб.")
    sheet.cell(row=40, column=4, value=100)
    sheet.cell(row=40, column=5, value=120)
    sheet.cell(row=40, column=6, value=90)
    sheet.cell(row=40, column=7, value=-30)

    sheet.cell(row=53, column=3, value="итого 2026")
    sheet.cell(row=53, column=4, value="ПЛАН")
    sheet.cell(row=53, column=5, value="ПРОГНОЗ")
    sheet.cell(row=53, column=6, value="ФАКТ+\nактуализ.ПРОГНОЗ")
    sheet.cell(row=53, column=7, value="ФАКТ")
    sheet.cell(row=53, column=8, value="Остаток к продаже")
    sheet.cell(row=54, column=3, value="Продажи, руб.")
    sheet.cell(row=54, column=4, value=2000)
    sheet.cell(row=54, column=5, value=2100)
    sheet.cell(row=54, column=6, value=2200)
    sheet.cell(row=54, column=7, value=500)
    sheet.cell(row=54, column=8, value=1700)

    sheet.cell(row=68, column=3, value="ИТОГО ПРОЕКТ")
    sheet.cell(row=68, column=4, value="ПЛАН")
    sheet.cell(row=68, column=5, value="ПРОГНОЗ")
    sheet.cell(row=68, column=6, value="ФАКТ+\nактуализ.ПРОГНОЗ")
    sheet.cell(row=68, column=7, value="ФАКТ")
    sheet.cell(row=68, column=8, value="Остаток к продаже")
    sheet.cell(row=69, column=3, value="Поступление ден. средств, руб.")
    sheet.cell(row=69, column=4, value=3000)
    sheet.cell(row=69, column=5, value=3100)
    sheet.cell(row=69, column=6, value=3200)
    sheet.cell(row=69, column=7, value=700)
    sheet.cell(row=69, column=8, value=2500)
    workbook.save(file_path)

    source, facts = parse_sales_plan_execution_file(file_path, "obvodny")

    assert source.snapshot_month == date(2026, 4, 1)
    assert source.snapshot_date == date(2026, 4, 30)
    assert len(facts) == 25

    project_sales = [fact for fact in facts if fact.block_kind == "segment_cumulative" and fact.segment == "project_total" and fact.metric_key == "sales_revenue"]
    assert {fact.owner_scope for fact in project_sales} == {"all", "developer"}
    assert {fact.scenario for fact in project_sales if fact.owner_scope == "all"} == {"plan", "fact", "deviation", "fact_forecast", "forecast_deviation"}

    month_sales = [fact for fact in facts if fact.block_kind == "month" and fact.metric_key == "sales_revenue"]
    assert {fact.scenario for fact in month_sales} == {"plan", "forecast", "fact", "fact_minus_forecast"}
    assert all(fact.period_month == date(2026, 4, 1) for fact in month_sales)

    year_sales = [fact for fact in facts if fact.block_kind == "year" and fact.metric_key == "sales_revenue"]
    assert {fact.scenario for fact in year_sales} == {"plan", "forecast", "fact_actualized_forecast", "fact", "remaining_to_sell"}
    assert all(fact.year == 2026 for fact in year_sales)

    lifetime_receipts = [fact for fact in facts if fact.block_kind == "project_lifetime" and fact.metric_key == "cash_receipts"]
    assert len(lifetime_receipts) == 5
    assert {fact.unit for fact in lifetime_receipts} == {"rub"}


def test_import_sales_plan_execution_is_idempotent(tmp_path) -> None:
    file_path = tmp_path / "Отчет об исполнении плана продаж (Обв.118)_31.03.2026.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.cell(row=2, column=4, value="Отчет об исполнении плана продаж проекта Обводный 118 на 31.03.2026")
    sheet.cell(row=4, column=3, value="ИТОГО ПРОЕКТ")
    sheet.cell(row=4, column=4, value="ПЛАН")
    sheet.cell(row=4, column=5, value="ФАКТ")
    sheet.cell(row=4, column=6, value="Отклонение")
    sheet.cell(row=4, column=8, value="ФАКТ+ПРОГНОЗ")
    sheet.cell(row=4, column=9, value="Отклонение")
    sheet.cell(row=5, column=3, value="Продажи, руб.")
    sheet.cell(row=5, column=4, value=1000)
    sheet.cell(row=5, column=5, value=900)
    sheet.cell(row=5, column=6, value=-100)
    sheet.cell(row=5, column=8, value=1100)
    sheet.cell(row=5, column=9, value=100)
    workbook.save(file_path)

    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)

    with Session(engine) as session:
        first = import_sales_plan_execution(session, tmp_path, "obvodny")
        second = import_sales_plan_execution(session, tmp_path, "obvodny")

    with Session(engine) as session:
        assert first.files == 1
        assert first.sources == 1
        assert first.facts == 5
        assert second.facts == 5
        assert session.query(SalesPlanExecutionSource).count() == 1
        assert session.query(SalesPlanExecutionFact).count() == 5
        fact = session.query(SalesPlanExecutionFact).filter(SalesPlanExecutionFact.scenario == "fact").one()
        assert fact.metric_key == "sales_revenue"
        assert fact.value == 900
