from datetime import date
from pathlib import Path

from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from app.db.models import Base, SalesReportFact, SalesReportSource
from app.importers.sales_report import import_sales_report, parse_sales_report_file, parse_snapshot_month_from_filename


def test_parse_snapshot_month_from_filename() -> None:
    period = parse_snapshot_month_from_filename(
        Path("\u041e\u0442\u0447\u0435\u0442 \u043e \u043f\u0440\u043e\u0434\u0430\u0436\u0430\u0445 (\u041e\u0431\u0432.118) Stories 30.04.2026.xlsx"),
    )

    assert period == date(2026, 4, 1)


def test_parse_sales_report_file_reads_segments_metrics_and_months(tmp_path) -> None:
    file_path = tmp_path / "\u041e\u0442\u0447\u0435\u0442 \u043e \u043f\u0440\u043e\u0434\u0430\u0436\u0430\u0445 (\u041e\u0431\u0432.118) Stories 30.04.2026.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "\u041b\u0438\u0441\u04421"
    sheet.cell(row=2, column=8, value=date(2026, 4, 30))
    sheet.cell(row=4, column=2, value="\u041e\u043f\u0438\u0441\u0430\u043d\u0438\u0435")
    sheet.cell(row=4, column=3, value="\u0418\u0442\u043e\u0433\u043e")
    sheet.cell(row=4, column=4, value="\u0424\u0430\u043a\u0442")
    sheet.cell(row=5, column=4, value=date(2026, 4, 1))
    sheet.cell(row=4, column=5, value="\u041f\u043b\u0430\u043d")
    sheet.cell(row=5, column=5, value=date(2026, 5, 1))
    sheet.cell(row=7, column=2, value="\u0418\u0442\u043e\u0433\u043e \u043f\u043e \u043f\u0440\u043e\u0435\u043a\u0442\u0443")
    sheet.cell(row=8, column=2, value="\u0412\u044b\u0440\u0443\u0447\u043a\u0430 \u043f\u043e \u043a\u043e\u043d\u0442\u0440\u0430\u043a\u0442\u0430\u0446\u0438\u0438")
    sheet.cell(row=8, column=3, value=1000)
    sheet.cell(row=8, column=4, value=100)
    sheet.cell(row=8, column=5, value=200)
    sheet.cell(row=9, column=2, value="     \u0417\u0430\u0441\u0442\u0440\u043e\u0439\u0449\u0438\u043a")
    sheet.cell(row=9, column=3, value=800)
    sheet.cell(row=9, column=4, value=80)
    sheet.cell(row=10, column=2, value="     \u0412\u0435\u043b\u043b")
    sheet.cell(row=10, column=3, value=200)
    sheet.cell(row=10, column=4, value=20)
    sheet.cell(row=20, column=2, value="\u0410\u043f\u0430\u0440\u0442\u0430\u043c\u0435\u043d\u0442\u044b")
    sheet.cell(row=24, column=2, value="\u041e\u0431\u044a\u0435\u043c \u043a\u043e\u043d\u0442\u0440\u0430\u043a\u0442\u0430\u0446\u0438\u0438, \u043a\u0432.\u043c.")
    sheet.cell(row=24, column=3, value=300)
    sheet.cell(row=24, column=4, value=30)
    sheet.cell(row=26, column=2, value="     \u0432 \u0442.\u0447. \u0412\u0435\u043b\u043b")
    sheet.cell(row=26, column=3, value=70)
    sheet.cell(row=26, column=4, value=7)
    workbook.save(file_path)

    source, facts = parse_sales_report_file(file_path, "obvodny")

    assert source.snapshot_month == date(2026, 4, 1)
    assert source.snapshot_date == date(2026, 4, 30)
    assert len(facts) == 11

    total_revenue = [fact for fact in facts if fact.segment == "project_total" and fact.metric_key == "contract_revenue"]
    assert {fact.owner_scope for fact in total_revenue} == {"all", "developer", "well"}
    assert {fact.period_kind for fact in total_revenue} == {"total", "month"}
    assert any(fact.period_month == date(2026, 4, 1) and fact.scenario == "fact" and fact.value == 100 for fact in total_revenue)
    assert any(fact.period_month == date(2026, 5, 1) and fact.scenario == "plan" and fact.value == 200 for fact in total_revenue)

    apartment_area = [fact for fact in facts if fact.segment == "apartments" and fact.metric_key == "contract_area_sqm"]
    assert {fact.owner_scope for fact in apartment_area} == {"all", "well_including"}
    assert all(fact.unit == "sqm" for fact in apartment_area)


def test_import_sales_report_is_idempotent(tmp_path) -> None:
    file_path = tmp_path / "\u041e\u0442\u0447\u0435\u0442 \u043e \u043f\u0440\u043e\u0434\u0430\u0436\u0430\u0445 (\u041e\u0431\u0432.118) Stories 31.03.2026.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.cell(row=2, column=8, value=date(2026, 3, 31))
    sheet.cell(row=4, column=3, value="\u0418\u0442\u043e\u0433\u043e")
    sheet.cell(row=4, column=4, value="\u0424\u0430\u043a\u0442")
    sheet.cell(row=5, column=4, value=date(2026, 3, 1))
    sheet.cell(row=7, column=2, value="\u0418\u0442\u043e\u0433\u043e \u043f\u043e \u043f\u0440\u043e\u0435\u043a\u0442\u0443")
    sheet.cell(row=8, column=2, value="\u0412\u044b\u0440\u0443\u0447\u043a\u0430 \u043f\u043e \u043a\u043e\u043d\u0442\u0440\u0430\u043a\u0442\u0430\u0446\u0438\u0438")
    sheet.cell(row=8, column=3, value=1000)
    sheet.cell(row=8, column=4, value=100)
    workbook.save(file_path)

    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)

    with Session(engine) as session:
        first = import_sales_report(session, tmp_path, "obvodny")
        second = import_sales_report(session, tmp_path, "obvodny")

    with Session(engine) as session:
        assert first.files == 1
        assert first.sources == 1
        assert first.facts == 2
        assert second.facts == 2
        assert session.query(SalesReportSource).count() == 1
        assert session.query(SalesReportFact).count() == 2
        fact = session.query(SalesReportFact).filter(SalesReportFact.period_kind == "month").one()
        assert fact.segment == "project_total"
        assert fact.metric_key == "contract_revenue"
        assert fact.scenario == "fact"
