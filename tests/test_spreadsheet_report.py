from io import BytesIO

from openpyxl import load_workbook

from app.pipeline.calculation_engine import CalculationResult
from app.pipeline.spreadsheet_report import build_model_raw_xlsx_report, should_send_xlsx_report
from app.pipeline.response_data import ResponseData


def test_build_model_raw_xlsx_report_returns_openable_table() -> None:
    calculation = CalculationResult(
        kind="sql_result",
        rows=[
            {
                "raw_sheet": "Финмодель",
                "row_number": 5,
                "row_label": "Площадь участка, га",
                "visible_cells": 3,
                "values_preview": "B: Площадь участка, га | E: 1.8641 | M: Валовая рентабельность",
            },
        ],
        columns=["raw_sheet", "row_number", "row_label", "visible_cells", "values_preview"],
        row_count=1,
        metrics=[],
    )

    file_bytes, filename = build_model_raw_xlsx_report(calculation)

    workbook = load_workbook(BytesIO(file_bytes))
    worksheet = workbook.active
    assert filename == "model.xlsx"
    assert worksheet.cell(row=1, column=1).value == "Лист"
    assert worksheet.cell(row=1, column=5).value == "B"
    assert worksheet.cell(row=1, column=6).value == "E"
    assert worksheet.cell(row=2, column=3).value == "Площадь участка, га"
    assert worksheet.cell(row=2, column=6).value == "1.8641"


def test_should_send_xlsx_report_only_for_model_raw_rows() -> None:
    calculation = CalculationResult(kind="sql_result", rows=[], columns=[], row_count=31, metrics=[])
    response_data = ResponseData(
        ready=True,
        title="",
        summary=[],
        table=None,
        source={"report_type": "model", "view": "model_raw_rows"},
        warnings=[],
        errors=[],
    )

    assert should_send_xlsx_report(response_data, calculation) is True

    response_data.source["view"] = "model_summary"
    assert should_send_xlsx_report(response_data, calculation) is False

    response_data.source["view"] = "model_raw_rows"
    calculation.row_count = 30
    assert should_send_xlsx_report(response_data, calculation) is False
