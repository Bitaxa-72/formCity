from datetime import date
from pathlib import Path

from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from app.db.models import AgentsReportDeal, AgentsReportMonthlyValue, AgentsReportSource, Base
from app.importers.agents_report import import_agents_report, parse_agents_report_file, parse_snapshot_date_from_filename


def test_parse_snapshot_date_from_filename() -> None:
    snapshot_date = parse_snapshot_date_from_filename(Path("Отчет по Агентам_по 30.04.2026 включительно.xlsx"))

    assert snapshot_date == date(2026, 4, 30)


def test_parse_agents_report_file_reads_deals_and_monthly_values(tmp_path) -> None:
    file_path = tmp_path / "Отчет по Агентам_по 30.04.2026 включительно.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Лист1"
    sheet.cell(row=3, column=1, value="Наименование Агента\n(юр.лицо, не брендовое название)")
    sheet.cell(row=3, column=2, value="№ помещ.")
    sheet.cell(row=3, column=3, value="ФИО")
    sheet.cell(row=3, column=4, value="№ ДДУ")
    sheet.cell(row=3, column=5, value="Дата")
    sheet.cell(row=3, column=6, value="Площадь")
    sheet.cell(row=3, column=7, value="Цена сделки для расчета Агентского вознаграждения, руб.")
    sheet.cell(row=3, column=8, value="Проверка\nQ-W \n(ДДУ+Уступка-Меблировка)")
    sheet.cell(row=3, column=9, value="Проверка\nG-H")
    sheet.cell(row=3, column=10, value="Размер Агентского вознаграждения\n%")
    sheet.cell(row=3, column=11, value="Размер Агентского вознаграждения\nруб.")
    sheet.cell(row=3, column=12, value="ИТОГО сумма по Акту")
    sheet.cell(row=3, column=13, value="Оплачено")
    sheet.cell(row=3, column=14, value="Остаток к оплате")
    sheet.cell(row=3, column=15, value="№ и дата Акта")
    sheet.cell(row=3, column=16, value="В бюджет месяц")
    sheet.cell(row=3, column=17, value="ДДУ+УСТУПКА\nруб.")
    sheet.cell(row=3, column=18, value="ДДУ+УСТУПКА\nруб./м2")
    sheet.cell(row=3, column=19, value="ДДУ\nруб.")
    sheet.cell(row=3, column=20, value="ДДУ\nруб./м2")
    sheet.cell(row=3, column=21, value="УСТУПКА\nруб.")
    sheet.cell(row=3, column=22, value="УСТУПКА\nруб./м2")
    sheet.cell(row=3, column=23, value="ДС Меблировка, руб")
    sheet.cell(row=3, column=24, value="Примечание")
    sheet.cell(row=3, column=27, value="ПРОШЛЫЕ ПЕРИОДЫ")
    sheet.cell(row=3, column=28, value=date(2026, 1, 1))
    sheet.cell(row=3, column=41, value="ПРОШЛЫЕ ПЕРИОДЫ")
    sheet.cell(row=3, column=42, value=date(2026, 1, 1))

    sheet.cell(row=4, column=1, value="ИП Тестовый Агент")
    sheet.cell(row=4, column=2, value="4.46")
    sheet.cell(row=4, column=3, value="Иванов Иван Иванович")
    sheet.cell(row=4, column=4, value="В-4046-ОБВ_ДУ")
    sheet.cell(row=4, column=5, value=date(2026, 1, 23))
    sheet.cell(row=4, column=6, value=26.93)
    sheet.cell(row=4, column=7, value=9056200)
    sheet.cell(row=4, column=8, value=9056200)
    sheet.cell(row=4, column=9, value=0)
    sheet.cell(row=4, column=10, value=0.06)
    sheet.cell(row=4, column=11, value=543372)
    sheet.cell(row=4, column=12, value=543372)
    sheet.cell(row=4, column=13, value=100000)
    sheet.cell(row=4, column=14, value=443372)
    sheet.cell(row=4, column=15, value="№ 1 от 01.04.2026")
    sheet.cell(row=4, column=16, value=date(2026, 5, 1))
    sheet.cell(row=4, column=17, value=9056200)
    sheet.cell(row=4, column=18, value=336286.6691)
    sheet.cell(row=4, column=19, value=8079000)
    sheet.cell(row=4, column=20, value=300000)
    sheet.cell(row=4, column=21, value=977200)
    sheet.cell(row=4, column=22, value=36286.6691)
    sheet.cell(row=4, column=23, value=0)
    sheet.cell(row=4, column=24, value="можно закрывать")
    sheet.cell(row=4, column=27, value=8079000)
    sheet.cell(row=4, column=28, value=1000)
    sheet.cell(row=4, column=41, value=977200)
    sheet.cell(row=4, column=42, value=2000)
    sheet.cell(row=5, column=1, value="ИТОГО:")
    sheet.cell(row=5, column=12, value=543372)
    workbook.save(file_path)

    source, deals, monthly_values = parse_agents_report_file(file_path, "obvodny")

    assert source.snapshot_date == date(2026, 4, 30)
    assert source.snapshot_month == date(2026, 4, 1)
    assert len(deals) == 1
    assert len(monthly_values) == 4

    deal = deals[0]
    assert deal.agent_name == "ИП Тестовый Агент"
    assert deal.buyer_name == "Иванов Иван Иванович"
    assert deal.ddu_number == "В-4046-ОБВ_ДУ"
    assert deal.commission_amount == 543372
    assert deal.ddu_amount == 8079000
    assert deal.assignment_amount == 977200
    assert deal.is_sensitive is True
    assert deal.sensitive_fields == {"agent_name": True, "buyer_name": True, "ddu_number": True, "act_info": True}

    assert {(value.value_kind, value.period_kind, value.period_month, value.value) for value in monthly_values} == {
        ("ddu_schedule", "past_periods_total", None, 8079000),
        ("ddu_schedule", "month", date(2026, 1, 1), 1000),
        ("assignment_schedule", "past_periods_total", None, 977200),
        ("assignment_schedule", "month", date(2026, 1, 1), 2000),
    }


def test_import_agents_report_is_idempotent(tmp_path) -> None:
    file_path = tmp_path / "Отчет по Агентам_на 24.03.2026.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.cell(row=3, column=1, value="Наименование Агента\n(юр.лицо, не брендовое название)")
    sheet.cell(row=3, column=2, value="№ помещ.")
    sheet.cell(row=3, column=3, value="ФИО")
    sheet.cell(row=3, column=4, value="№ ДДУ")
    sheet.cell(row=3, column=5, value="Дата")
    sheet.cell(row=3, column=6, value="Площадь")
    sheet.cell(row=3, column=7, value="Цена сделки для расчета Агентского вознаграждения, руб.")
    sheet.cell(row=3, column=8, value="Размер Агентского вознаграждения\n%")
    sheet.cell(row=3, column=9, value="Размер Агентского вознаграждения\nруб.")
    sheet.cell(row=4, column=1, value="ИП Тестовый Агент")
    sheet.cell(row=4, column=2, value="4.46")
    sheet.cell(row=4, column=3, value="Иванов Иван Иванович")
    sheet.cell(row=4, column=4, value="В-4046-ОБВ_ДУ")
    sheet.cell(row=4, column=5, value=date(2026, 1, 23))
    sheet.cell(row=4, column=6, value=26.93)
    sheet.cell(row=4, column=7, value=9056200)
    sheet.cell(row=4, column=8, value=0.06)
    sheet.cell(row=4, column=9, value=543372)
    workbook.save(file_path)

    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)

    with Session(engine) as session:
        first = import_agents_report(session, tmp_path, "obvodny")
        second = import_agents_report(session, tmp_path, "obvodny")

    with Session(engine) as session:
        assert first.files == 1
        assert first.sources == 1
        assert first.deals == 1
        assert second.deals == 1
        assert session.query(AgentsReportSource).count() == 1
        assert session.query(AgentsReportDeal).count() == 1
        assert session.query(AgentsReportMonthlyValue).count() == 0
