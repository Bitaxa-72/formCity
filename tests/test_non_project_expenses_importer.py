from datetime import date
from pathlib import Path

from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from app.db.models import Base, NonProjectExpenseFact, NonProjectExpenseSource
from app.importers.non_project_expenses import import_non_project_expenses, parse_non_project_file, parse_period_month_from_filename


def test_parse_period_month_from_filename() -> None:
    period = parse_period_month_from_filename(
        Path("\u041d\u0435\u043f\u0440\u043e\u0435\u043a\u0442\u043d\u044b\u0435 \u0440\u0430\u0441\u0445\u043e\u0434\u044b_(30.04.2026).xlsx"),
    )

    assert period == date(2026, 4, 1)


def test_parse_non_project_file_reads_detail_and_summary_rows(tmp_path) -> None:
    file_path = tmp_path / "\u041d\u0435\u043f\u0440\u043e\u0435\u043a\u0442\u043d\u044b\u0435 \u0440\u0430\u0441\u0445\u043e\u0434\u044b_(30.04.2026).xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "\u041b\u0438\u0441\u04421"
    sheet.cell(row=1, column=2, value="\u0434\u0430\u0442\u0430 \u0437\u0430\u043f\u043e\u043b\u043d\u0435\u043d\u0438\u044f")
    sheet.cell(row=1, column=3, value=date(2026, 4, 30))
    sheet.cell(row=2, column=2, value="\u0432 \u0424\u041c")
    sheet.cell(row=2, column=3, value="\u041d\u0430\u0438\u043c\u0435\u043d\u043e\u0432\u0430\u043d\u0438\u0435")
    sheet.cell(row=2, column=4, value="\u0421\u0443\u043c\u043c\u0430, \u0440\u0443\u0431.")
    sheet.cell(row=2, column=5, value="\u0418\u0441\u043f\u043e\u043b\u043d\u0435\u043d\u043e")
    sheet.cell(row=2, column=6, value="\u041e\u0441\u0442\u0430\u0442\u043e\u043a/\u043f\u0440\u043e\u0433\u043d\u043e\u0437")
    sheet.cell(row=2, column=7, value="\u0421\u043f\u0440\u0430\u0432\u043e\u0447\u043d\u043e")
    sheet.cell(row=3, column=2, value="\u043a\u0432\u0430\u0440\u0442\u0438\u0440\u044b, \u0430\u043f\u0430\u0440\u0442\u044b")
    sheet.cell(row=3, column=3, value="\u041d\u0435\u0434\u043e\u043f\u043e\u043b\u0443\u0447\u0435\u043d\u043d\u044b\u0439 \u0434\u043e\u0445\u043e\u0434")
    sheet.cell(row=3, column=4, value=100)
    sheet.cell(row=3, column=5, value=60)
    sheet.cell(row=3, column=6, value=40)
    sheet.cell(row=3, column=7, value="\u0421\u043f\u0440\u0430\u0432\u043a\u0430")
    sheet.cell(row=4, column=3, value="\u041d\u0435\u043f\u0440\u043e\u0435\u043a\u0442\u043d\u044b\u0435 \u0440\u0430\u0441\u0445\u043e\u0434\u044b")
    sheet.cell(row=4, column=4, value=200)
    sheet.cell(row=4, column=5, value=150)
    sheet.cell(row=4, column=6, value=50)
    workbook.save(file_path)

    source, facts = parse_non_project_file(file_path, "all")

    assert source.period_month == date(2026, 4, 1)
    assert source.filled_at == date(2026, 4, 30)
    assert len(facts) == 2
    assert facts[0].row_type == "detail"
    assert facts[0].item_kind == "lost_income"
    assert facts[0].amount == 100
    assert facts[0].executed_amount == 60
    assert facts[0].remaining_amount == 40
    assert facts[1].row_type == "summary"
    assert facts[1].item_kind == "non_project_expenses_total"


def test_import_non_project_expenses_is_idempotent(tmp_path) -> None:
    file_path = tmp_path / "\u041d\u0435\u043f\u0440\u043e\u0435\u043a\u0442\u043d\u044b\u0435 \u0440\u0430\u0441\u0445\u043e\u0434\u044b_(30.04.2026).xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.cell(row=1, column=3, value=date(2026, 4, 30))
    sheet.cell(row=3, column=2, value="\u041f\u0418\u0420")
    sheet.cell(row=3, column=3, value="\u041f\u0440\u043e\u0435\u043a\u0442\u043d\u044b\u0435 \u0440\u0430\u0431\u043e\u0442\u044b")
    sheet.cell(row=3, column=4, value=300)
    workbook.save(file_path)

    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)

    with Session(engine) as session:
        first = import_non_project_expenses(session, tmp_path, "all")
        second = import_non_project_expenses(session, tmp_path, "all")

    with Session(engine) as session:
        assert first.files == 1
        assert first.sources == 1
        assert first.facts == 1
        assert second.facts == 1
        assert session.query(NonProjectExpenseSource).count() == 1
        assert session.query(NonProjectExpenseFact).count() == 1
        assert session.query(NonProjectExpenseFact).one().item_kind == "pir"
