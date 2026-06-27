from datetime import date

from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from app.db.models import Base, ModelKpiFact, ModelMonthlyFact, ModelRawCell, ModelRawRow, ModelRawSheet, ModelSource
from app.importers.model import import_model


def test_import_model_reads_sources_monthly_and_kpi(tmp_path) -> None:
    source = tmp_path / "model"
    source.mkdir()
    file_path = source / "Модель fact 04.26.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ФМ_"
    sheet.cell(row=3, column=7, value=date(2026, 4, 1))
    sheet.cell(row=4, column=7, value="факт")
    sheet.cell(row=7, column=2, value="1")
    sheet.cell(row=7, column=3, value="Выручка")
    sheet.cell(row=7, column=7, value=100)

    plan_sheet = workbook.create_sheet("ФМ_ПЛАН")
    plan_sheet.cell(row=3, column=7, value=date(2026, 4, 1))
    plan_sheet.cell(row=7, column=2, value="1")
    plan_sheet.cell(row=7, column=3, value="Выручка")
    plan_sheet.cell(row=7, column=7, value=90)

    kpi_sheet = workbook.create_sheet("NEWKPI's_")
    kpi_sheet.cell(row=3, column=5, value="руб.")
    kpi_sheet.cell(row=4, column=5, value="NPV")
    kpi_sheet.cell(row=4, column=6, value=50)

    raw_sheet = workbook.create_sheet("Финмодель")
    raw_sheet.cell(row=1, column=1, value="Проект:")
    raw_sheet.cell(row=1, column=2, value="Обводный")
    raw_sheet.cell(row=2, column=1, value="Выручка")
    raw_sheet.cell(row=2, column=2, value=1000)
    workbook.save(file_path)

    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)

    with Session(engine) as session:
        result = import_model(session, source, "obvodny")

    with Session(engine) as session:
        assert result.files == 1
        assert session.query(ModelSource).count() == 1
        assert session.query(ModelMonthlyFact).count() == 2
        assert session.query(ModelKpiFact).count() == 1
        assert result.raw_sheets == 1
        assert result.raw_rows == 2
        assert result.raw_cells == 4
        assert session.query(ModelRawSheet).count() == 1
        assert session.query(ModelRawRow).count() == 2
        assert session.query(ModelRawCell).count() == 4
        assert session.query(ModelMonthlyFact).filter_by(scenario="current").one().value == 100
        assert session.query(ModelMonthlyFact).filter_by(scenario="current").one().metric_key == "model_revenue"
        assert session.query(ModelKpiFact).one().metric_key == "model_npv"
        assert session.query(ModelRawCell).filter_by(sheet_name="Финмодель", row_number=2, column_number=2).one().value_number == 1000
