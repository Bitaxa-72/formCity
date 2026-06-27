from datetime import date
from pathlib import Path

from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from app.db.models import Base, SummaryCell, SummaryRow, SummarySheet, SummarySource
from app.importers.summary import import_summary, infer_project, parse_summary_file


def build_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Апартаменты"
    sheet.cell(row=1, column=1, value="ЖК Тест")
    sheet.cell(row=2, column=2, value="Проектные данные")
    sheet.cell(row=3, column=1, value="№п/п")
    sheet.cell(row=3, column=2, value="ФИО Клиента")
    sheet.cell(row=3, column=3, value="№ апарт.")
    sheet.cell(row=3, column=4, value="Этаж")
    sheet.cell(row=3, column=5, value="Площадь")
    sheet.cell(row=3, column=6, value="Цена ДДУ")
    sheet.cell(row=3, column=7, value="Дата ДДУ")
    sheet.cell(row=4, column=1, value=1)
    sheet.cell(row=4, column=2, value="Иванов Иван")
    sheet.cell(row=4, column=3, value="4.12")
    sheet.cell(row=4, column=4, value=4)
    sheet.cell(row=4, column=5, value=42.5)
    sheet.cell(row=4, column=6, value=12000000)
    sheet.cell(row=4, column=7, value=date(2026, 5, 1))

    totals = workbook.create_sheet("Сводная")
    totals.cell(row=3, column=1, value="Показатель")
    totals.cell(row=3, column=2, value="Площадь")
    totals.cell(row=3, column=3, value="Выручка")
    totals.cell(row=4, column=1, value="ИТОГО")
    totals.cell(row=4, column=2, value=42.5)
    totals.cell(row=4, column=3, value=12000000)
    workbook.save(path)


def test_infer_project() -> None:
    source = Path("сводная")

    assert infer_project(source / "велл московский" / "Сводная_Велл Московский.xlsx", source) == "moskovsky"
    assert infer_project(source / "обводный" / "Сводная_Обводный 118.xlsx", source) == "obvodny"
    assert infer_project(source / "евгеньевский" / "ЖК Евгеньевский.xlsx", source) == "evgenievsky"


def test_parse_summary_file_reads_rows_cells_and_sensitive_marks(tmp_path) -> None:
    source = tmp_path / "сводная"
    folder = source / "обводный"
    folder.mkdir(parents=True)
    file_path = folder / "Сводная_Обводный 118.xlsx"
    build_workbook(file_path)

    parsed_source, sheets, rows, cells = parse_summary_file(file_path, source)

    assert parsed_source.project == "obvodny"
    assert len(sheets) == 2
    assert len(rows) == 6
    assert len(cells) == 22

    apartments_sheet = next(sheet for sheet in sheets if sheet.sheet_name == "Апартаменты")
    assert apartments_sheet.sheet_kind == "residential_units"
    assert apartments_sheet.header_row == 3

    detail_row = next(row for row in rows if row.sheet_name == "Апартаменты" and row.row_number == 4)
    assert detail_row.row_type == "detail"
    assert detail_row.customer_name == "Иванов Иван"
    assert detail_row.unit_number == "4.12"
    assert detail_row.is_sensitive is True
    assert detail_row.raw_values["ФИО Клиента"] == "Иванов Иван"

    price_cell = next(cell for cell in cells if cell.sheet_name == "Апартаменты" and cell.row_number == 4 and cell.header_key == "цена_дду")
    assert price_cell.value_type == "number"
    assert price_cell.value_number == 12000000
    assert price_cell.is_sensitive is False

    area_cell = next(cell for cell in cells if cell.sheet_name == "Сводная" and cell.row_number == 4 and cell.header_key == "площадь")
    assert area_cell.value_number == 42.5
    assert area_cell.is_sensitive is False


def test_import_summary_is_idempotent(tmp_path) -> None:
    source = tmp_path / "сводная"
    folder = source / "евгеньевский"
    folder.mkdir(parents=True)
    file_path = folder / "ЖК Евгеньевский.xlsx"
    build_workbook(file_path)

    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)

    with Session(engine) as session:
        first = import_summary(session, source)
        second = import_summary(session, source)

    with Session(engine) as session:
        assert first.files == 1
        assert first.sources == 1
        assert first.sheets == 2
        assert second.sheets == 2
        assert session.query(SummarySource).count() == 1
        assert session.query(SummarySheet).count() == 2
        assert session.query(SummaryRow).count() == 6
        assert session.query(SummaryCell).count() == 22
