import argparse
import hashlib
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import delete
from sqlalchemy.orm import Session

from app.core.config import load_settings
from app.db.database import create_database_tables, get_session_factory
from app.db.models import NonProjectExpenseFact, NonProjectExpenseSource
from app.pipeline.sensitive_data import detect_sensitive_kind


DEFAULT_SOURCE = Path("..") / "\u043e\u0440\u0438\u0433\u0438\u043d\u0430\u043b\u044b \u0442\u0430\u0431\u043b\u0438\u0446" / "\u043d\u0435 \u043f\u0440\u043e\u0435\u043a\u0442\u043d\u044b\u0435 \u0440\u0430\u0441\u0445\u043e\u0434\u044b"
DEFAULT_PROJECT = "all"
SOURCE_DATE_RE = re.compile(r"(\d{2})\.(\d{2})\.(\d{4})")
ASCII_SOURCE_DATE_RE = re.compile(r"(\d{4})-(\d{2})-(\d{2})")


@dataclass(frozen=True)
class NonProjectExpenseImportResult:
    files: int
    sources: int
    facts: int


def normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    text = " ".join(str(value).strip().split())
    return text or None


def normalize_search_text(value: str | None) -> str:
    return " ".join((value or "").strip().lower().replace("\u0451", "\u0435").split())


def parse_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int | float):
        return Decimal(str(value))
    if isinstance(value, str):
        prepared = value.strip().replace(" ", "").replace(",", ".")
        if not prepared:
            return None
        try:
            return Decimal(prepared)
        except InvalidOperation:
            return None
    return None


def month_start(value: date) -> date:
    return date(value.year, value.month, 1)


def parse_date_value(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value[:19]).date()
        except ValueError:
            return None
    return None


def parse_period_month_from_filename(path: Path) -> date | None:
    match = SOURCE_DATE_RE.search(path.name)
    if match:
        return date(int(match.group(3)), int(match.group(2)), 1)

    match = ASCII_SOURCE_DATE_RE.search(path.name)
    if match:
        return date(int(match.group(1)), int(match.group(2)), 1)

    return None


def file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def classify_item_kind(fm_category: str | None, item_name: str) -> str:
    category = normalize_search_text(fm_category)
    name = normalize_search_text(item_name)
    source = category or name

    if "квартиры" in source or "апарты" in source or "недополуч" in name or "упущенн" in name or "продажа ресторана" in name:
        return "lost_income"
    if name.startswith("дз "):
        return "debt_receivable"
    if "непроектные расходы" in name:
        return "non_project_expenses_total"
    if name == "личное":
        return "personal"
    if name == "ахр":
        return "admin_expenses"
    if name == "евг":
        return "evgenievsky"
    if name.startswith("ооо"):
        return "legal_entity"
    if "отделоч" in source:
        return "fit_out"
    if "коммерческ" in source:
        return "commercial"
    if "мебел" in source:
        return "furniture"
    if "строитель" in source:
        return "construction"
    if "содержание застройщика" in source:
        return "developer_maintenance"
    if "содержание объекта" in source:
        return "object_maintenance"
    if "финансов" in source:
        return "finance"
    if "пир" in source:
        return "pir"
    if "прочие" in source:
        return "other_income_expense"
    return "other"


def row_sensitive_kind(*values: Any) -> str | None:
    text = " ".join(str(value) for value in values if value is not None)
    return detect_sensitive_kind(text)


def parse_non_project_file(path: Path, project: str) -> tuple[NonProjectExpenseSource, list[NonProjectExpenseFact]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    filled_at = parse_date_value(sheet.cell(row=1, column=3).value)
    period_month = parse_period_month_from_filename(path) or (month_start(filled_at) if filled_at else None)
    if period_month is None:
        raise ValueError("non_project_expense_period_not_found")

    source = NonProjectExpenseSource(
        project=project,
        period_month=period_month,
        filled_at=filled_at,
        file_name=path.name,
        file_hash=file_hash(path),
    )

    facts = []
    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row_index <= 2:
            continue

        fm_category = normalize_text(row[1] if len(row) > 1 else None)
        item_name = normalize_text(row[2] if len(row) > 2 else None)
        amount = parse_decimal(row[3] if len(row) > 3 else None)
        executed_amount = parse_decimal(row[4] if len(row) > 4 else None)
        remaining_amount = parse_decimal(row[5] if len(row) > 5 else None)
        reference_text = normalize_text(row[6] if len(row) > 6 else None)
        if not item_name or amount is None:
            continue

        sensitive_kind = row_sensitive_kind(fm_category, item_name, reference_text)
        facts.append(
            NonProjectExpenseFact(
                project=project,
                period_month=period_month,
                filled_at=filled_at,
                row_order=row_index,
                row_type="detail" if fm_category else "summary",
                item_kind=classify_item_kind(fm_category, item_name),
                fm_category=fm_category,
                item_name=item_name,
                amount=amount,
                executed_amount=executed_amount,
                remaining_amount=remaining_amount,
                reference_text=reference_text,
                unit="rub",
                is_sensitive=sensitive_kind is not None,
                sensitive_kind=sensitive_kind,
                source_sheet=sheet.title,
                source_row=row_index,
                source_file=path.name,
            ),
        )

    return source, facts


def find_xlsx_files(source: Path) -> list[Path]:
    return sorted(path for path in source.rglob("*.xlsx") if not path.name.startswith("~$"))


def replace_non_project_expenses(
    session: Session,
    project: str,
    sources: list[NonProjectExpenseSource],
    facts: list[NonProjectExpenseFact],
) -> int:
    periods = {source.period_month for source in sources}
    if periods:
        session.execute(
            delete(NonProjectExpenseFact).where(
                NonProjectExpenseFact.project == project,
                NonProjectExpenseFact.period_month.in_(periods),
            ),
        )
        session.execute(
            delete(NonProjectExpenseSource).where(
                NonProjectExpenseSource.project == project,
                NonProjectExpenseSource.period_month.in_(periods),
            ),
        )

    session.add_all(sources)
    session.add_all(facts)
    session.commit()
    return len(facts)


def import_non_project_expenses(session: Session, source: Path, project: str) -> NonProjectExpenseImportResult:
    files = find_xlsx_files(source)
    sources = []
    facts = []
    for file_path in files:
        parsed_source, parsed_facts = parse_non_project_file(file_path, project)
        sources.append(parsed_source)
        facts.extend(parsed_facts)

    fact_count = replace_non_project_expenses(session, project, sources, facts)
    return NonProjectExpenseImportResult(files=len(files), sources=len(sources), facts=fact_count)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--project", default=DEFAULT_PROJECT)
    args = parser.parse_args()

    settings = load_settings()
    create_database_tables(settings.database_url)
    session_factory = get_session_factory(settings.database_url)
    with session_factory() as session:
        result = import_non_project_expenses(session, args.source, args.project)

    print(f"Imported {result.files} files, {result.facts} non-project expense facts")


if __name__ == "__main__":
    main()
