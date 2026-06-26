import argparse
import calendar
import hashlib
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import delete
from sqlalchemy.orm import Session

from app.core.config import load_settings
from app.db.database import create_database_tables, get_session_factory
from app.db.models import (
    DebtBookingDeviation,
    DebtBookingItem,
    DebtBookingMonthlyValue,
    DebtBookingRefusal,
    DebtBookingSource,
)


DEFAULT_SOURCE = Path("..") / "оригиналы таблиц" / "дз и брони"
DEFAULT_PROJECT = "obvodny"
SOURCE_DATE_RE = re.compile(r"(\d{2})\.(\d{2})\.(\d{4})")
MONTH_YEAR_RE = re.compile(r"(январ[ьяе]|феврал[ьяе]|март[ае]?|апрел[ьяе]|ма[йяе]|июн[ьяе]|июл[ьяе]|август[ае]?|сентябр[ьяе]|октябр[ьяе]|ноябр[ьяе]|декабр[ьяе])\s+(\d{4})", re.IGNORECASE)
RU_MONTHS = {
    "январ": 1,
    "феврал": 2,
    "март": 3,
    "апрел": 4,
    "ма": 5,
    "июн": 6,
    "июл": 7,
    "август": 8,
    "сентябр": 9,
    "октябр": 10,
    "ноябр": 11,
    "декабр": 12,
}
CATEGORY_KINDS = {
    "итого": "total",
    "зарегистрировано": "registered",
    "просроченные": "overdue",
    "текущие": "current",
    "дупт подписан, не зарегистрирован": "dupt_signed_unregistered",
    "брони": "booking",
}


@dataclass(frozen=True)
class DebtBookingsImportResult:
    files: int
    sources: int
    items: int
    monthly_values: int
    deviations: int
    refusals: int


def normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    text = " ".join(str(value).strip().split())
    return text or None


def normalize_search_text(value: Any) -> str:
    return " ".join((str(value) if value is not None else "").strip().lower().replace("ё", "е").split())


def parse_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int | float):
        return Decimal(str(value))
    if isinstance(value, str):
        prepared = value.strip().replace(" ", "").replace("\xa0", "").replace(",", ".")
        if not prepared:
            return None
        try:
            return Decimal(prepared)
        except InvalidOperation:
            return None
    return None


def parse_date_value(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        match = SOURCE_DATE_RE.search(value)
        if match:
            return date(int(match.group(3)), int(match.group(2)), int(match.group(1)))
        try:
            return datetime.fromisoformat(value.strip()[:19]).date()
        except ValueError:
            return None
    return None


def month_start(value: date) -> date:
    return date(value.year, value.month, 1)


def month_end(year: int, month: int) -> date:
    return date(year, month, calendar.monthrange(year, month)[1])


def parse_month_year_text(value: Any) -> date | None:
    if value is None:
        return None
    match = MONTH_YEAR_RE.search(normalize_search_text(value))
    if match is None:
        return None
    month_text = match.group(1)
    year = int(match.group(2))
    for prefix, month in RU_MONTHS.items():
        if month_text.startswith(prefix):
            return date(year, month, 1)
    return None


def parse_snapshot_date_from_filename(path: Path) -> date | None:
    match = SOURCE_DATE_RE.search(path.name)
    if match:
        return date(int(match.group(3)), int(match.group(2)), int(match.group(1)))
    period_month = parse_month_year_text(path.name)
    if period_month is None:
        return None
    return month_end(period_month.year, period_month.month)


def file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def find_xlsx_files(source: Path) -> list[Path]:
    return sorted(path for path in source.rglob("*.xlsx") if not path.name.startswith("~$"))


def value_at(rows: list[tuple[Any, ...]], row: int, column: int) -> Any:
    row_values = rows[row - 1]
    index = column - 1
    if index < 0 or index >= len(row_values):
        return None
    return row_values[index]


def item_kind_for(label: str | None) -> str | None:
    normalized = normalize_search_text(label)
    if normalized.endswith(":"):
        normalized = normalized[:-1]
    return CATEGORY_KINDS.get(normalized)


def is_category_row(label: str | None, unit_number: str | None) -> bool:
    return item_kind_for(label) is not None and unit_number is None


def find_main_sheet(workbook) -> Any:
    for sheet in workbook.worksheets:
        normalized = normalize_search_text(sheet.title)
        if "дз" in normalized or "брон" in normalized:
            return sheet
    return workbook.active


def find_sheet(workbook, title_part: str) -> Any | None:
    title_part = normalize_search_text(title_part)
    for sheet in workbook.worksheets:
        if title_part in normalize_search_text(sheet.title):
            return sheet
    return None


def parse_snapshot_date_from_sheet(rows: list[tuple[Any, ...]]) -> date | None:
    for row in rows[:8]:
        for value in row:
            parsed = parse_date_value(value)
            if parsed is not None:
                return parsed
    return None


def find_month_columns(headers: tuple[Any, ...]) -> list[tuple[int, date]]:
    month_columns = []
    for column, value in enumerate(headers, 1):
        parsed = parse_date_value(value)
        if parsed is not None:
            month_columns.append((column, month_start(parsed)))
    return month_columns


def parse_main_sheet(
    rows: list[tuple[Any, ...]],
    sheet_title: str,
    source_file: str,
    project: str,
    snapshot_date: date,
) -> tuple[list[DebtBookingItem], list[DebtBookingMonthlyValue]]:
    snapshot_month = month_start(snapshot_date)
    headers = rows[2]
    month_columns = find_month_columns(headers)
    items = []
    monthly_values = []
    current_section: str | None = None
    current_kind = "detail"

    for row in range(4, len(rows) + 1):
        label = normalize_text(value_at(rows, row, 6))
        manager_name = normalize_text(value_at(rows, row, 7))
        special_marker = normalize_text(value_at(rows, row, 8))
        unit_number = normalize_text(value_at(rows, row, 9))
        total_amount = parse_decimal(value_at(rows, row, 10))
        comments = normalize_text(value_at(rows, row, 23))
        contacts = normalize_text(value_at(rows, row, 24))
        month_values = [(column, period_month, parse_decimal(value_at(rows, row, column))) for column, period_month in month_columns]
        has_month_value = any(value is not None for _, _, value in month_values)

        if not any((label, manager_name, unit_number, total_amount, comments, contacts, has_month_value)):
            continue

        category_kind = item_kind_for(label)
        row_type = "category" if is_category_row(label, unit_number) else "detail"
        if row_type == "category":
            current_section = label
            current_kind = category_kind or current_kind
            item_kind = category_kind or current_kind
            section = label
        else:
            item_kind = current_kind if current_kind != "total" else "detail"
            section = current_section

        item = DebtBookingItem(
            project=project,
            snapshot_month=snapshot_month,
            snapshot_date=snapshot_date,
            row_order=row,
            row_type=row_type,
            item_kind=item_kind,
            section=section,
            client_name=label,
            manager_name=manager_name,
            is_special_client=normalize_search_text(special_marker) == "+",
            unit_number=unit_number,
            total_amount=total_amount,
            comments=comments,
            contacts=contacts,
            unit="rub",
            is_sensitive=True,
            sensitive_fields={
                "client_name": True,
                "manager_name": True,
                "unit_number": True,
                "comments": True,
                "contacts": True,
            },
            source_sheet=sheet_title,
            source_row=row,
            source_file=source_file,
        )
        items.append(item)

        for column, period_month, value in month_values:
            if value is None:
                continue
            monthly_values.append(
                DebtBookingMonthlyValue(
                    project=project,
                    snapshot_month=snapshot_month,
                    snapshot_date=snapshot_date,
                    item_source_row=row,
                    item_kind=item_kind,
                    row_type=row_type,
                    period_month=period_month,
                    value=value,
                    unit="rub",
                    source_sheet=sheet_title,
                    source_row=row,
                    source_col=column,
                    source_file=source_file,
                ),
            )

    return items, monthly_values


def parse_deviation_sheet(
    rows: list[tuple[Any, ...]],
    sheet_title: str,
    source_file: str,
    project: str,
    snapshot_date: date,
) -> list[DebtBookingDeviation]:
    snapshot_month = month_start(snapshot_date)
    period_month = None
    if len(rows) >= 2:
        for value in rows[1]:
            period_month = parse_month_year_text(value)
            if period_month is not None:
                break

    deviations = []
    current_section: str | None = None
    current_kind = "detail"

    for row in range(3, len(rows) + 1):
        label = normalize_text(value_at(rows, row, 2))
        unit_number = normalize_text(value_at(rows, row, 3))
        plan_amount = parse_decimal(value_at(rows, row, 4))
        updated_plan_amount = parse_decimal(value_at(rows, row, 5))
        plan_comment = normalize_text(value_at(rows, row, 6))
        fact_payment_amount = parse_decimal(value_at(rows, row, 8))
        remaining_amount = parse_decimal(value_at(rows, row, 9))
        fact_comment = normalize_text(value_at(rows, row, 10))

        if not any((label, unit_number, plan_amount, updated_plan_amount, plan_comment, fact_payment_amount, remaining_amount, fact_comment)):
            continue

        category_kind = item_kind_for(label)
        row_type = "category" if is_category_row(label, unit_number) else "detail"
        if row_type == "category":
            current_section = label
            current_kind = category_kind or current_kind
            item_kind = category_kind or current_kind
            section = label
        else:
            item_kind = current_kind if current_kind != "total" else "detail"
            section = current_section

        deviations.append(
            DebtBookingDeviation(
                project=project,
                snapshot_month=snapshot_month,
                snapshot_date=snapshot_date,
                period_month=period_month,
                row_order=row,
                row_type=row_type,
                item_kind=item_kind,
                section=section,
                client_name=label,
                unit_number=unit_number,
                plan_amount=plan_amount,
                updated_plan_amount=updated_plan_amount,
                plan_comment=plan_comment,
                fact_payment_amount=fact_payment_amount,
                remaining_amount=remaining_amount,
                fact_comment=fact_comment,
                unit="rub",
                is_sensitive=True,
                sensitive_fields={
                    "client_name": True,
                    "unit_number": True,
                    "plan_comment": True,
                    "fact_comment": True,
                },
                source_sheet=sheet_title,
                source_row=row,
                source_file=source_file,
            ),
        )

    return deviations


def parse_refusal_sheet(
    rows: list[tuple[Any, ...]],
    sheet_title: str,
    source_file: str,
    project: str,
    snapshot_date: date,
) -> list[DebtBookingRefusal]:
    snapshot_month = month_start(snapshot_date)
    refusals = []
    for row in range(2, len(rows) + 1):
        customer_name = normalize_text(value_at(rows, row, 1))
        status = normalize_text(value_at(rows, row, 2))
        area_sqm = parse_decimal(value_at(rows, row, 3))
        unit_number = normalize_text(value_at(rows, row, 4))
        full_price_amount = parse_decimal(value_at(rows, row, 5))
        payment_type = normalize_text(value_at(rows, row, 6))
        reason = normalize_text(value_at(rows, row, 7))
        agency = normalize_text(value_at(rows, row, 8))
        manager_name = normalize_text(value_at(rows, row, 9))

        if not any((customer_name, status, area_sqm, unit_number, full_price_amount, payment_type, reason, agency, manager_name)):
            continue

        refusals.append(
            DebtBookingRefusal(
                project=project,
                snapshot_month=snapshot_month,
                snapshot_date=snapshot_date,
                row_order=row,
                customer_name=customer_name,
                status=status,
                area_sqm=area_sqm,
                unit_number=unit_number,
                full_price_amount=full_price_amount,
                payment_type=payment_type,
                reason=reason,
                agency=agency,
                manager_name=manager_name,
                unit="rub",
                is_sensitive=True,
                sensitive_fields={
                    "customer_name": True,
                    "unit_number": True,
                    "reason": True,
                    "manager_name": True,
                },
                source_sheet=sheet_title,
                source_row=row,
                source_file=source_file,
            ),
        )

    return refusals


def parse_debt_bookings_file(
    path: Path,
    project: str,
) -> tuple[DebtBookingSource, list[DebtBookingItem], list[DebtBookingMonthlyValue], list[DebtBookingDeviation], list[DebtBookingRefusal]]:
    workbook = load_workbook(path, data_only=True, read_only=True, keep_links=False)
    main_sheet = find_main_sheet(workbook)
    main_rows = list(main_sheet.iter_rows(values_only=True))
    snapshot_date = parse_snapshot_date_from_sheet(main_rows) or parse_snapshot_date_from_filename(path)
    if snapshot_date is None:
        raise ValueError("debt_bookings_period_not_found")
    snapshot_month = month_start(snapshot_date)

    source = DebtBookingSource(
        project=project,
        snapshot_month=snapshot_month,
        snapshot_date=snapshot_date,
        file_name=path.name,
        file_hash=file_hash(path),
    )

    items, monthly_values = parse_main_sheet(main_rows, main_sheet.title, path.name, project, snapshot_date)

    deviation_sheet = find_sheet(workbook, "Отклонения")
    deviations = []
    if deviation_sheet is not None:
        deviations = parse_deviation_sheet(
            list(deviation_sheet.iter_rows(values_only=True)),
            deviation_sheet.title,
            path.name,
            project,
            snapshot_date,
        )

    refusal_sheet = find_sheet(workbook, "Отказы")
    refusals = []
    if refusal_sheet is not None:
        refusals = parse_refusal_sheet(
            list(refusal_sheet.iter_rows(values_only=True)),
            refusal_sheet.title,
            path.name,
            project,
            snapshot_date,
        )

    return source, items, monthly_values, deviations, refusals


def replace_debt_bookings(
    session: Session,
    project: str,
    sources: list[DebtBookingSource],
    items: list[DebtBookingItem],
    monthly_values: list[DebtBookingMonthlyValue],
    deviations: list[DebtBookingDeviation],
    refusals: list[DebtBookingRefusal],
) -> tuple[int, int, int, int]:
    dates = {source.snapshot_date for source in sources}
    if dates:
        for model in (
            DebtBookingMonthlyValue,
            DebtBookingDeviation,
            DebtBookingRefusal,
            DebtBookingItem,
            DebtBookingSource,
        ):
            session.execute(
                delete(model).where(
                    model.project == project,
                    model.snapshot_date.in_(dates),
                ),
            )

    session.add_all(sources)
    session.add_all(items)
    session.add_all(monthly_values)
    session.add_all(deviations)
    session.add_all(refusals)
    session.commit()
    return len(items), len(monthly_values), len(deviations), len(refusals)


def import_debt_bookings(session: Session, source: Path, project: str) -> DebtBookingsImportResult:
    files = find_xlsx_files(source)
    sources = []
    items = []
    monthly_values = []
    deviations = []
    refusals = []
    for file_path in files:
        parsed_source, parsed_items, parsed_monthly_values, parsed_deviations, parsed_refusals = parse_debt_bookings_file(file_path, project)
        sources.append(parsed_source)
        items.extend(parsed_items)
        monthly_values.extend(parsed_monthly_values)
        deviations.extend(parsed_deviations)
        refusals.extend(parsed_refusals)

    item_count, monthly_value_count, deviation_count, refusal_count = replace_debt_bookings(
        session,
        project,
        sources,
        items,
        monthly_values,
        deviations,
        refusals,
    )
    return DebtBookingsImportResult(
        files=len(files),
        sources=len(sources),
        items=item_count,
        monthly_values=monthly_value_count,
        deviations=deviation_count,
        refusals=refusal_count,
    )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--project", default=DEFAULT_PROJECT)
    args = parser.parse_args()

    settings = load_settings()
    create_database_tables(settings.database_url)
    session_factory = get_session_factory(settings.database_url)
    with session_factory() as session:
        result = import_debt_bookings(session, args.source, args.project)

    print(
        f"Imported {result.files} files, {result.items} debt booking items, "
        f"{result.monthly_values} monthly values, {result.deviations} deviations, {result.refusals} refusals",
    )


if __name__ == "__main__":
    main()
