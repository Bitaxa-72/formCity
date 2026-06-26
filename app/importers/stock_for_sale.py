import argparse
import hashlib
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import delete
from sqlalchemy.orm import Session

from app.core.config import load_settings
from app.db.database import create_database_tables, get_session_factory
from app.db.models import StockForSaleFact, StockForSaleSource


DEFAULT_SOURCE = Path("..") / "\u043e\u0440\u0438\u0433\u0438\u043d\u0430\u043b\u044b \u0442\u0430\u0431\u043b\u0438\u0446" / "\u043e\u0441\u0442\u0430\u0442\u043a\u0438 \u0432 \u043f\u0440\u043e\u0434\u0430\u0436\u0435"
DEFAULT_PROJECT = "obvodny"
SOURCE_DATE_RE = re.compile(r"(\d{2})\.(\d{2})\.(\d{4})")
ASCII_SOURCE_DATE_RE = re.compile(r"(\d{4})-(\d{2})-(\d{2})")
FLOOR_RE = re.compile(r"(\d+)\s*\u044d\u0442\u0430\u0436")


@dataclass(frozen=True)
class StockForSaleImportResult:
    files: int
    sources: int
    facts: int


def normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    text = " ".join(str(value).strip().split())
    return text or None


def normalize_search_text(value: str | None) -> str:
    return " ".join((value or "").strip().lower().replace("\u0451", "\u0435").split())


def parse_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int | float):
        return Decimal(str(value))
    if isinstance(value, str):
        prepared = value.strip().replace(" ", "").replace(",", ".")
        if not prepared:
            return None
        try:
            return Decimal(prepared)
        except InvalidOperation:
            return None
    return None


def parse_int(value: Any) -> int | None:
    decimal = parse_decimal(value)
    if decimal is None:
        return None
    return int(decimal)


def parse_date_value(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value[:19]).date()
        except ValueError:
            return None
    return None


def month_start(value: date) -> date:
    return date(value.year, value.month, 1)


def parse_snapshot_month_from_filename(path: Path) -> date | None:
    match = SOURCE_DATE_RE.search(path.name)
    if match:
        return date(int(match.group(3)), int(match.group(2)), 1)

    match = ASCII_SOURCE_DATE_RE.search(path.name)
    if match:
        return date(int(match.group(1)), int(match.group(2)), 1)

    return None


def file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def add_decimal(left: Decimal | None, right: Decimal | None) -> Decimal | None:
    if left is None and right is None:
        return None
    return (left or Decimal("0")) + (right or Decimal("0"))


def classify_row(label: str | None, ddu_amount: Decimal | None, row_order: int) -> tuple[str, str, str, int | None, bool]:
    text = normalize_search_text(label)
    floor_match = FLOOR_RE.search(text)
    floor_number = int(floor_match.group(1)) if floor_match else None
    is_in_work = "\u0432 \u0440\u0430\u0431\u043e\u0442\u0435" in text

    if label is None and ddu_amount is not None:
        return "total_with_markup", "\u0412\u0441\u0435\u0433\u043e \u0441 \u043d\u0430\u0446\u0435\u043d\u043a\u043e\u0439 \u0414\u0423\u041f\u0422", "total", None, False
    if text == "\u0432 \u0442\u043e\u043c \u0447\u0438\u0441\u043b\u0435:":
        return "section", label or "", "section", None, False
    if text == "\u0432\u0441\u0435\u0433\u043e":
        return "total", label or "", "total", None, False
    if "\u043a\u043b\u0430\u0434\u043e\u0432\u043a" in text:
        return "category", label or "", "storage", floor_number, is_in_work
    if "\u0440\u0435\u0441\u0442\u043e\u0440\u0430\u043d" in text:
        return "category", label or "", "restaurant", floor_number, is_in_work
    if "\u0441\u0437 \u043e\u0431\u0432" in text:
        return "detail", label or "", "developer_balance", floor_number, is_in_work
    if "\u0430\u043f\u0430\u0440\u0442" in text:
        return "category" if not is_in_work else "detail", label or "", "apartment", floor_number, is_in_work
    if floor_number == 1:
        return "detail", label or "", "first_floor", floor_number, is_in_work
    if floor_number is not None:
        return "detail", label or "", "apartment", floor_number, is_in_work
    return "detail", label or f"row_{row_order}", "other", floor_number, is_in_work


def parse_stock_for_sale_file(path: Path, project: str) -> tuple[StockForSaleSource, list[StockForSaleFact]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    snapshot_date = parse_date_value(sheet.cell(row=4, column=2).value)
    snapshot_month = parse_snapshot_month_from_filename(path) or (month_start(snapshot_date) if snapshot_date else None)
    if snapshot_month is None:
        raise ValueError("stock_for_sale_period_not_found")

    source = StockForSaleSource(
        project=project,
        snapshot_month=snapshot_month,
        snapshot_date=snapshot_date,
        file_name=path.name,
        file_hash=file_hash(path),
    )

    facts = []
    for row_order, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row_order <= 5:
            continue

        label = normalize_text(row[1] if len(row) > 1 else None)
        ddu_amount = parse_decimal(row[2] if len(row) > 2 else None)
        dupt_markup_amount = parse_decimal(row[3] if len(row) > 3 else None)
        area_sqm = parse_decimal(row[4] if len(row) > 4 else None)
        unit_count = parse_int(row[5] if len(row) > 5 else None)
        ddu_price_per_sqm = parse_decimal(row[6] if len(row) > 6 else None)
        dupt_price_per_sqm = parse_decimal(row[7] if len(row) > 7 else None)

        if label is None and ddu_amount is None:
            continue

        row_type, row_label, property_type, floor_number, is_in_work = classify_row(label, ddu_amount, row_order)
        if row_type == "section":
            continue

        if row_type == "total_with_markup":
            total_amount = ddu_amount
            ddu_amount = None
        else:
            total_amount = add_decimal(ddu_amount, dupt_markup_amount)

        total_price_per_sqm = None
        if total_amount is not None and area_sqm not in (None, Decimal("0")):
            total_price_per_sqm = total_amount / area_sqm

        facts.append(
            StockForSaleFact(
                project=project,
                snapshot_month=snapshot_month,
                snapshot_date=snapshot_date,
                row_order=row_order,
                row_type=row_type,
                row_label=row_label,
                property_type=property_type,
                floor_number=floor_number,
                is_in_work=is_in_work,
                ddu_amount=ddu_amount,
                dupt_markup_amount=dupt_markup_amount,
                total_amount=total_amount,
                area_sqm=area_sqm,
                unit_count=unit_count,
                ddu_price_per_sqm=ddu_price_per_sqm,
                dupt_price_per_sqm=dupt_price_per_sqm,
                total_price_per_sqm=total_price_per_sqm,
                unit="mixed",
                source_sheet=sheet.title,
                source_row=row_order,
                source_file=path.name,
            ),
        )

    return source, facts


def find_xlsx_files(source: Path) -> list[Path]:
    return sorted(path for path in source.rglob("*.xlsx") if not path.name.startswith("~$"))


def replace_stock_for_sale(
    session: Session,
    project: str,
    sources: list[StockForSaleSource],
    facts: list[StockForSaleFact],
) -> int:
    periods = {source.snapshot_month for source in sources}
    if periods:
        session.execute(
            delete(StockForSaleFact).where(
                StockForSaleFact.project == project,
                StockForSaleFact.snapshot_month.in_(periods),
            ),
        )
        session.execute(
            delete(StockForSaleSource).where(
                StockForSaleSource.project == project,
                StockForSaleSource.snapshot_month.in_(periods),
            ),
        )

    session.add_all(sources)
    session.add_all(facts)
    session.commit()
    return len(facts)


def import_stock_for_sale(session: Session, source: Path, project: str) -> StockForSaleImportResult:
    files = find_xlsx_files(source)
    sources = []
    facts = []
    for file_path in files:
        parsed_source, parsed_facts = parse_stock_for_sale_file(file_path, project)
        sources.append(parsed_source)
        facts.extend(parsed_facts)

    fact_count = replace_stock_for_sale(session, project, sources, facts)
    return StockForSaleImportResult(files=len(files), sources=len(sources), facts=fact_count)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--project", default=DEFAULT_PROJECT)
    args = parser.parse_args()

    settings = load_settings()
    create_database_tables(settings.database_url)
    session_factory = get_session_factory(settings.database_url)
    with session_factory() as session:
        result = import_stock_for_sale(session, args.source, args.project)

    print(f"Imported {result.files} files, {result.facts} stock-for-sale facts")


if __name__ == "__main__":
    main()
