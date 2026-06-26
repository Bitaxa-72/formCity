import argparse
import hashlib
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import delete
from sqlalchemy.orm import Session

from app.core.config import load_settings
from app.db.database import create_database_tables, get_session_factory
from app.db.models import SalesReportFact, SalesReportSource


DEFAULT_SOURCE = Path("..") / "\u043e\u0440\u0438\u0433\u0438\u043d\u0430\u043b\u044b \u0442\u0430\u0431\u043b\u0438\u0446" / "\u043e\u0442\u0447\u0435\u0442 \u043e \u043f\u0440\u043e\u0434\u0430\u0436\u0430\u0445"
DEFAULT_PROJECT = "obvodny"
SOURCE_DATE_RE = re.compile(r"(\d{2})\.(\d{2})\.(\d{4})")
ASCII_SOURCE_DATE_RE = re.compile(r"(\d{4})-(\d{2})-(\d{2})")


SEGMENT_BY_LABEL = {
    "\u0438\u0442\u043e\u0433\u043e \u043f\u043e \u043f\u0440\u043e\u0435\u043a\u0442\u0443": ("project_total", "\u0418\u0442\u043e\u0433\u043e \u043f\u043e \u043f\u0440\u043e\u0435\u043a\u0442\u0443"),
    "\u0430\u043f\u0430\u0440\u0442\u0430\u043c\u0435\u043d\u0442\u044b": ("apartments", "\u0410\u043f\u0430\u0440\u0442\u0430\u043c\u0435\u043d\u0442\u044b"),
    "\u043a\u043e\u043c\u043c\u0435\u0440\u0446\u0438\u044f 1 \u044d\u0442\u0430\u0436": ("commercial_1_floor", "\u041a\u043e\u043c\u043c\u0435\u0440\u0446\u0438\u044f 1 \u044d\u0442\u0430\u0436"),
    "\u0440\u0435\u0441\u0442\u043e\u0440\u0430\u043d": ("restaurant", "\u0420\u0435\u0441\u0442\u043e\u0440\u0430\u043d"),
    "\u043a\u043b\u0430\u0434\u043e\u0432\u043a\u0438": ("storage", "\u041a\u043b\u0430\u0434\u043e\u0432\u043a\u0438"),
    "\u043a\u043e\u043c\u043c\u0435\u0440\u0446\u0438\u044f 2 \u044d\u0442\u0430\u0436": ("commercial_2_floor", "\u041a\u043e\u043c\u043c\u0435\u0440\u0446\u0438\u044f 2 \u044d\u0442\u0430\u0436"),
    "sh": ("sh", "SH"),
}


@dataclass(frozen=True)
class SalesReportImportResult:
    files: int
    sources: int
    facts: int


def normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    text = " ".join(str(value).strip().split())
    return text or None


def normalize_search_text(value: str | None) -> str:
    return " ".join((value or "").strip().lower().replace("\u0451", "\u0435").split())


def parse_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int | float):
        return Decimal(str(value))
    if isinstance(value, str):
        prepared = value.strip().replace(" ", "").replace(",", ".")
        if not prepared:
            return None
        try:
            return Decimal(prepared)
        except InvalidOperation:
            return None
    return None


def parse_date_value(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value[:19]).date()
        except ValueError:
            return None
    return None


def month_start(value: date) -> date:
    return date(value.year, value.month, 1)


def parse_snapshot_month_from_filename(path: Path) -> date | None:
    match = SOURCE_DATE_RE.search(path.name)
    if match:
        return date(int(match.group(3)), int(match.group(2)), 1)

    match = ASCII_SOURCE_DATE_RE.search(path.name)
    if match:
        return date(int(match.group(1)), int(match.group(2)), 1)

    return None


def file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def metric_key_for(label: str) -> tuple[str, str, str] | None:
    normalized = normalize_search_text(label)
    if normalized == "\u0432\u044b\u0440\u0443\u0447\u043a\u0430 \u043f\u043e \u043a\u043e\u043d\u0442\u0440\u0430\u043a\u0442\u0430\u0446\u0438\u0438":
        return "contract_revenue", "\u0412\u044b\u0440\u0443\u0447\u043a\u0430 \u043f\u043e \u043a\u043e\u043d\u0442\u0440\u0430\u043a\u0442\u0430\u0446\u0438\u0438", "thousand_rub"
    if normalized == "\u0444\u0430\u043a\u0442\u0438\u0447\u0435\u0441\u043a\u0438\u0435 \u043e\u043f\u043b\u0430\u0442\u044b \u043f\u043e \u0434\u0434\u0443":
        return "ddu_actual_payments", "\u0424\u0430\u043a\u0442\u0438\u0447\u0435\u0441\u043a\u0438\u0435 \u043e\u043f\u043b\u0430\u0442\u044b \u043f\u043e \u0414\u0414\u0423", "thousand_rub"
    if normalized == "\u0433\u0440\u0430\u0444\u0438\u043a \u043e\u043f\u043b\u0430\u0442\u044b \u043e\u0441\u0442\u0430\u0442\u043a\u0430 \u043f\u043e \u0434\u0434\u0443":
        return "ddu_remaining_payment_schedule", "\u0413\u0440\u0430\u0444\u0438\u043a \u043e\u043f\u043b\u0430\u0442\u044b \u043e\u0441\u0442\u0430\u0442\u043a\u0430 \u043f\u043e \u0414\u0414\u0423", "thousand_rub"
    if normalized == "\u043e\u0431\u044a\u0435\u043c \u043a\u043e\u043d\u0442\u0440\u0430\u043a\u0442\u0430\u0446\u0438\u0438, \u043a\u0432.\u043c.":
        return "contract_area_sqm", "\u041e\u0431\u044a\u0435\u043c \u043a\u043e\u043d\u0442\u0440\u0430\u043a\u0442\u0430\u0446\u0438\u0438, \u043a\u0432.\u043c.", "sqm"
    if normalized == "\u043e\u0431\u044a\u0435\u043c \u043a\u043e\u043d\u0442\u0440\u0430\u043a\u0442\u0430\u0446\u0438\u0438, \u0448\u0442., \u0432 \u0442.\u0447.:":
        return "contract_count", "\u041e\u0431\u044a\u0435\u043c \u043a\u043e\u043d\u0442\u0440\u0430\u043a\u0442\u0430\u0446\u0438\u0438, \u0448\u0442.", "count"
    if normalized == "\u0446\u0435\u043d\u0430 \u0437\u0430 1 \u043a\u0432.\u043c.":
        return "price_per_sqm", "\u0426\u0435\u043d\u0430 \u0437\u0430 1 \u043a\u0432.\u043c.", "thousand_rub_per_sqm"
    if normalized == "\u043d\u0430\u043a\u043e\u043f\u0438\u0442\u0435\u043b\u044c\u043d\u0430\u044f \u0446\u0435\u043d\u0430 \u0437\u0430 1 \u043a\u0432.\u043c.":
        return "cumulative_price_per_sqm", "\u041d\u0430\u043a\u043e\u043f\u0438\u0442\u0435\u043b\u044c\u043d\u0430\u044f \u0446\u0435\u043d\u0430 \u0437\u0430 1 \u043a\u0432.\u043c.", "thousand_rub_per_sqm"
    return None


def owner_scope_for(label: str) -> str:
    normalized = normalize_search_text(label)
    if normalized == "\u0437\u0430\u0441\u0442\u0440\u043e\u0439\u0449\u0438\u043a":
        return "developer"
    if normalized == "\u0432\u0435\u043b\u043b":
        return "well"
    if normalized == "\u0432 \u0442.\u0447. \u0432\u0435\u043b\u043b":
        return "well_including"
    return "all"


def find_xlsx_files(source: Path) -> list[Path]:
    return sorted(path for path in source.rglob("*.xlsx") if not path.name.startswith("~$"))


def parse_sales_report_file(path: Path, project: str) -> tuple[SalesReportSource, list[SalesReportFact]]:
    workbook = load_workbook(path, data_only=True, read_only=True, keep_links=False)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    def value_at(row: int, column: int) -> Any:
        row_index = row - 1
        column_index = column - 1
        if row_index < 0 or row_index >= len(rows):
            return None
        row_values = rows[row_index]
        if column_index < 0 or column_index >= len(row_values):
            return None
        return row_values[column_index]

    snapshot_date = parse_date_value(value_at(2, 8))
    snapshot_month = parse_snapshot_month_from_filename(path) or (month_start(snapshot_date) if snapshot_date else None)
    if snapshot_month is None:
        raise ValueError("sales_report_period_not_found")

    source = SalesReportSource(
        project=project,
        snapshot_month=snapshot_month,
        snapshot_date=snapshot_date,
        file_name=path.name,
        file_hash=file_hash(path),
    )

    period_columns: list[tuple[int, str, date | None, str]] = [(3, "total", None, "total")]
    max_column = max((len(row) for row in rows), default=0)
    for column in range(4, max_column + 1):
        scenario_text = normalize_search_text(value_at(4, column))
        period_month = parse_date_value(value_at(5, column))
        if scenario_text not in {"\u0444\u0430\u043a\u0442", "\u043f\u043b\u0430\u043d"} or period_month is None:
            continue
        scenario = "fact" if scenario_text == "\u0444\u0430\u043a\u0442" else "plan"
        period_columns.append((column, "month", month_start(period_month), scenario))

    facts = []
    current_segment: tuple[str, str] | None = None
    current_metric: tuple[str, str, str] | None = None
    for row in range(1, len(rows) + 1):
        label = normalize_text(value_at(row, 2))
        normalized = normalize_search_text(label)
        if normalized in SEGMENT_BY_LABEL:
            current_segment = SEGMENT_BY_LABEL[normalized]
            current_metric = None
            continue
        if current_segment is None or label is None:
            continue

        metric = metric_key_for(label)
        if metric is not None:
            current_metric = metric
            owner_scope = "all"
        elif current_metric is not None:
            owner_scope = owner_scope_for(label)
            if owner_scope == "all":
                continue
            metric = current_metric
        else:
            continue

        metric_key, metric_name, unit = metric
        segment, segment_label = current_segment
        for column, period_kind, period_month, scenario in period_columns:
            value = parse_decimal(value_at(row, column))
            if value is None:
                continue
            facts.append(
                SalesReportFact(
                    project=project,
                    snapshot_month=snapshot_month,
                    snapshot_date=snapshot_date,
                    segment=segment,
                    segment_label=segment_label,
                    metric_key=metric_key,
                    metric_name=metric_name,
                    owner_scope=owner_scope,
                    period_kind=period_kind,
                    period_month=period_month,
                    scenario=scenario,
                    value=value,
                    unit=unit,
                    source_sheet=sheet.title,
                    source_row=row,
                    source_col=column,
                    source_file=path.name,
                ),
            )

    return source, facts


def replace_sales_report(
    session: Session,
    project: str,
    sources: list[SalesReportSource],
    facts: list[SalesReportFact],
) -> int:
    periods = {source.snapshot_month for source in sources}
    if periods:
        session.execute(
            delete(SalesReportFact).where(
                SalesReportFact.project == project,
                SalesReportFact.snapshot_month.in_(periods),
            ),
        )
        session.execute(
            delete(SalesReportSource).where(
                SalesReportSource.project == project,
                SalesReportSource.snapshot_month.in_(periods),
            ),
        )

    session.add_all(sources)
    session.add_all(facts)
    session.commit()
    return len(facts)


def import_sales_report(session: Session, source: Path, project: str) -> SalesReportImportResult:
    files = find_xlsx_files(source)
    sources = []
    facts = []
    for file_path in files:
        parsed_source, parsed_facts = parse_sales_report_file(file_path, project)
        sources.append(parsed_source)
        facts.extend(parsed_facts)

    fact_count = replace_sales_report(session, project, sources, facts)
    return SalesReportImportResult(files=len(files), sources=len(sources), facts=fact_count)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--project", default=DEFAULT_PROJECT)
    args = parser.parse_args()

    settings = load_settings()
    create_database_tables(settings.database_url)
    session_factory = get_session_factory(settings.database_url)
    with session_factory() as session:
        result = import_sales_report(session, args.source, args.project)

    print(f"Imported {result.files} files, {result.facts} sales report facts")


if __name__ == "__main__":
    main()
