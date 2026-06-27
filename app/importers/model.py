import argparse
import hashlib
import re
import zipfile
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from time import time
from typing import Any
from xml.etree import ElementTree

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import delete
from sqlalchemy.orm import Session

from app.core.config import load_settings
from app.db.database import create_database_tables, get_session_factory
from app.db.models import (
    ModelAssumptionFact,
    ModelComparisonFact,
    ModelKpiFact,
    ModelMonthlyFact,
    ModelPassportFact,
    ModelRawCell,
    ModelRawRow,
    ModelRawSheet,
    ModelSource,
)
from app.importers.payment_calendar import excel_serial_to_month, read_xlsx_rows
from app.pipeline.sensitive_data import detect_sensitive_kind


DEFAULT_SOURCE = Path("..") / "\u043e\u0440\u0438\u0433\u0438\u043d\u0430\u043b\u044b \u0442\u0430\u0431\u043b\u0438\u0446" / "\u043c\u043e\u0434\u0435\u043b\u044c"
DEFAULT_PROJECT = "obvodny"
SHEET_FM = "\u0424\u041c_"
SHEET_FM_PLAN = "\u0424\u041c_\u041f\u041b\u0410\u041d"
SHEET_KPI = "NEWKPI's_"
SHEET_KPI_PLAN = "NEWKPI's_\u041f\u041b\u0410\u041d"
SHEET_COMPARISON = "\u0421\u0440\u0430\u0432\u043d\u0435\u043d\u0438\u0435"
SHEET_PASSPORT = "\u041f\u0430\u0441\u043f\u043e\u0440\u0442"
SHEET_RATES = "\u041f\u0440\u043e\u0446\u0435\u043d\u0442\u044b"
SHEET_CONSOLIDATION = "\u0414\u043b\u044f \u043a\u043e\u043d\u0441\u043e\u043b\u0438\u0434\u0430\u0446\u0438\u0438"
SHEET_FINMODEL = "\u0424\u0438\u043d\u043c\u043e\u0434\u0435\u043b\u044c"
SHEET_REMAINS = "\u041e\u0441\u0442\u0430\u0442\u043a\u0438"
RAW_MODEL_SHEETS = {
    SHEET_CONSOLIDATION: "consolidation",
    SHEET_FINMODEL: "financial_model",
    SHEET_REMAINS: "remains",
}
MODEL_FILE_MARKER = "\u041c\u043e\u0434\u0435\u043b\u044c"
SNAPSHOT_RE = re.compile(r"\b(\d{2})\.(\d{2})\b")
SECTION_CODE_RE = re.compile(r"^\d+\.?$")
MAIN_NAMESPACE = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
RELATIONSHIP_NAMESPACE = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
PACKAGE_RELATIONSHIP_NAMESPACE = "{http://schemas.openxmlformats.org/package/2006/relationships}"
METRIC_KEY_ALIASES = {
    "выручка": "model_revenue",
    "себестоимость продаж": "model_cost_of_sales",
    "валовая прибыль": "model_gross_profit",
    "чистая прибыль": "model_net_profit",
    "npv, млн.руб.": "model_npv",
    "npv": "model_npv",
    "roe, %": "model_roe",
    "roe": "model_roe",
    "llcr": "model_llcr",
    "общая площадь зданий, м2": "model_total_area",
    "общая площадь зданий": "model_total_area",
    "квартиры, шт. в т.ч.:": "model_units_count",
    "апарты, шт. в т.ч.:": "model_units_count",
    "пир": "model_pir",
}
SAFE_MODEL_METRIC_KEYS = set(METRIC_KEY_ALIASES.values())


@dataclass(frozen=True)
class ModelImportResult:
    files: int
    sources: int
    monthly: int
    kpi: int
    comparison: int
    passport: int
    assumptions: int
    raw_sheets: int = 0
    raw_rows: int = 0
    raw_cells: int = 0


def parse_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int | float):
        return Decimal(str(value))
    if isinstance(value, str):
        prepared = value.strip().replace(" ", "").replace(",", ".")
        if not prepared:
            return None
        try:
            return Decimal(prepared)
        except InvalidOperation:
            return None
    return None


def normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    text = " ".join(str(value).strip().split())
    return text or None


def normalize_metric_name(value: str | None) -> str:
    return " ".join((value or "").strip().lower().replace("ё", "е").split())


def resolve_metric_key(metric_name: str | None) -> str | None:
    return METRIC_KEY_ALIASES.get(normalize_metric_name(metric_name))


def parse_snapshot_month(path: Path) -> date:
    match = SNAPSHOT_RE.search(path.name)
    if not match:
        raise ValueError("model_snapshot_month_not_found")
    month = int(match.group(1))
    year = 2000 + int(match.group(2))
    return date(year, month, 1)


def file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def find_xlsx_files(source: Path) -> list[Path]:
    return sorted(
        path
        for path in source.rglob("*.xlsx")
        if not path.name.startswith("~$") and MODEL_FILE_MARKER.lower() in path.name.lower()
    )


def row_sensitive_kind(*values: Any) -> str | None:
    text = " ".join(value for value in values if isinstance(value, str) and value.strip())
    return detect_sensitive_kind(text)


def metric_sensitive_kind(metric_key: str | None, *values: Any) -> str | None:
    if metric_key in SAFE_MODEL_METRIC_KEYS:
        return None
    return row_sensitive_kind(*values)


def normalize_month(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date().replace(day=1)
    if isinstance(value, date):
        return value.replace(day=1)
    numeric = parse_decimal(value)
    if numeric is not None:
        try:
            return excel_serial_to_month(numeric)
        except (InvalidOperation, ValueError, OverflowError):
            return None
    return None


def read_sheet_paths(path: Path) -> dict[str, str]:
    with zipfile.ZipFile(path) as archive:
        workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
        rels = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))

    relation_targets = {
        relation.attrib["Id"]: relation.attrib["Target"]
        for relation in rels.findall(f"{PACKAGE_RELATIONSHIP_NAMESPACE}Relationship")
    }
    result = {}
    for sheet in workbook.findall(f"{MAIN_NAMESPACE}sheets/{MAIN_NAMESPACE}sheet"):
        sheet_name = sheet.attrib["name"]
        relation_id = sheet.attrib[f"{RELATIONSHIP_NAMESPACE}id"]
        target = relation_targets[relation_id].lstrip("/")
        result[sheet_name] = target if target.startswith("xl/") else f"xl/{target}"
    return result


def detect_date_columns(rows: dict[int, dict[int, str | None]]) -> list[tuple[int, date, str | None]]:
    date_row = rows.get(3, {})
    status_row = rows.get(4, {})
    result = []
    for column_index, value in sorted(date_row.items()):
        period_month = normalize_month(value)
        if period_month is None:
            continue
        status = normalize_text(status_row.get(column_index))
        result.append((column_index, period_month, status))
    return result


def build_monthly_facts(
    rows: dict[int, dict[int, str | None]],
    sheet_name: str,
    project: str,
    snapshot_month: date,
    scenario: str,
    source_file: str,
) -> list[ModelMonthlyFact]:
    return [ModelMonthlyFact(**row) for row in iter_monthly_fact_mappings(rows, sheet_name, project, snapshot_month, scenario, source_file)]


def iter_monthly_fact_mappings(
    rows: dict[int, dict[int, str | None]],
    sheet_name: str,
    project: str,
    snapshot_month: date,
    scenario: str,
    source_file: str,
    monthly_mode: str = "actual",
) -> Any:
    date_columns = detect_date_columns(rows)
    if not date_columns:
        return

    current_section: str | None = None
    for row_index in sorted(rows):
        if row_index < 7:
            continue
        row = rows[row_index]
        row_code = normalize_text(row.get(2))
        metric_name = normalize_text(row.get(3))
        if not metric_name:
            continue
        if row_code and SECTION_CODE_RE.fullmatch(row_code):
            current_section = metric_name

        metric_key = resolve_metric_key(metric_name)
        sensitive_kind = metric_sensitive_kind(metric_key, row_code, metric_name)
        for column_index, period_month, period_status in date_columns:
            if monthly_mode == "actual" and period_month > snapshot_month:
                continue
            value = parse_decimal(row.get(column_index))
            if value is None:
                continue
            yield {
                "project": project,
                "snapshot_month": snapshot_month,
                "scenario": scenario,
                "period_month": period_month,
                "period_status": period_status,
                "row_code": row_code,
                "section": current_section,
                "metric_name": metric_name,
                "metric_key": metric_key,
                "value": value,
                "normalized_value": value,
                "unit": "rub",
                "is_sensitive": sensitive_kind is not None,
                "sensitive_kind": sensitive_kind,
                "source_sheet": sheet_name,
                "source_row": row_index,
                "source_col": column_index + 1,
                "source_file": source_file,
            }


def build_kpi_facts(
    rows: dict[int, dict[int, str | None]],
    sheet_name: str,
    project: str,
    snapshot_month: date,
    scenario: str,
    source_file: str,
) -> list[ModelKpiFact]:
    facts = []
    current_section: str | None = None
    unit = normalize_text(rows.get(3, {}).get(5))
    for row_index in sorted(rows):
        row = rows[row_index]
        metric_name = normalize_text(row.get(5))
        value = parse_decimal(row.get(6))
        if not metric_name:
            continue
        if value is None:
            current_section = metric_name
            continue

        metric_key = resolve_metric_key(metric_name)
        sensitive_kind = metric_sensitive_kind(metric_key, metric_name)
        facts.append(
            ModelKpiFact(
                project=project,
                snapshot_month=snapshot_month,
                scenario=scenario,
                section=current_section,
                metric_name=metric_name,
                metric_key=metric_key,
                value=value,
                normalized_value=value,
                unit=unit,
                is_sensitive=sensitive_kind is not None,
                sensitive_kind=sensitive_kind,
                source_sheet=sheet_name,
                source_row=row_index,
                source_col=6,
                source_file=source_file,
            ),
        )
    return facts


def build_comparison_facts(
    rows: dict[int, dict[int, str | None]],
    sheet_name: str,
    project: str,
    snapshot_month: date,
    source_file: str,
) -> list[ModelComparisonFact]:
    facts = []
    for row_index in sorted(rows):
        row = rows[row_index]
        row_values = [row.get(column) for column in range(1, 13)]
        labels = [normalize_text(value) for value in row_values if normalize_text(value)]
        metric_name = next((label for label in labels if parse_decimal(label) is None), None)
        numeric_values = [parse_decimal(value) for value in row_values]
        numeric_values = [value for value in numeric_values if value is not None]
        if not metric_name or not numeric_values:
            continue

        metric_key = resolve_metric_key(metric_name)
        sensitive_kind = metric_sensitive_kind(metric_key, *row_values)
        facts.append(
            ModelComparisonFact(
                project=project,
                snapshot_month=snapshot_month,
                metric_name=metric_name,
                metric_key=metric_key,
                current_value=numeric_values[0] if len(numeric_values) > 0 else None,
                plan_value=numeric_values[1] if len(numeric_values) > 1 else None,
                deviation_value=numeric_values[2] if len(numeric_values) > 2 else None,
                deviation_percent=numeric_values[3] if len(numeric_values) > 3 else None,
                unit="rub",
                is_sensitive=sensitive_kind is not None,
                sensitive_kind=sensitive_kind,
                source_sheet=sheet_name,
                source_row=row_index,
                source_file=source_file,
            ),
        )
    return facts


def build_passport_facts(
    rows: dict[int, dict[int, str | None]],
    sheet_name: str,
    project: str,
    snapshot_month: date,
    source_file: str,
) -> list[ModelPassportFact]:
    facts = []
    current_section: str | None = None
    for row_index in sorted(rows):
        row = rows[row_index]
        values = [row.get(column) for column in range(1, 31)]
        text_values = [normalize_text(value) for value in values if normalize_text(value)]
        if not text_values:
            continue
        metric_name = text_values[0]
        if len(text_values) == 1:
            current_section = metric_name
            continue

        value_text = text_values[1]
        value_number = parse_decimal(value_text)
        metric_key = resolve_metric_key(metric_name)
        sensitive_kind = metric_sensitive_kind(metric_key, *text_values)
        facts.append(
            ModelPassportFact(
                project=project,
                snapshot_month=snapshot_month,
                section=current_section,
                metric_name=metric_name,
                metric_key=metric_key,
                value_text=value_text,
                value_number=value_number,
                is_sensitive=sensitive_kind is not None,
                sensitive_kind=sensitive_kind,
                source_sheet=sheet_name,
                source_row=row_index,
                source_col=2,
                source_file=source_file,
            ),
        )
    return facts


def build_assumption_facts(
    rows: dict[int, dict[int, str | None]],
    sheet_name: str,
    project: str,
    snapshot_month: date,
    source_file: str,
) -> list[ModelAssumptionFact]:
    facts = []
    for row_index in sorted(rows):
        row = rows[row_index]
        row_values = [row.get(column) for column in sorted(row)]
        metric_name = next((normalize_text(value) for value in row_values if normalize_text(value) and parse_decimal(value) is None), None)
        value = next((parse_decimal(value) for value in row_values if parse_decimal(value) is not None), None)
        if not metric_name or value is None:
            continue

        metric_key = resolve_metric_key(metric_name)
        sensitive_kind = metric_sensitive_kind(metric_key, *row_values)
        facts.append(
            ModelAssumptionFact(
                project=project,
                snapshot_month=snapshot_month,
                metric_name=metric_name,
                metric_key=metric_key,
                value=value,
                is_sensitive=sensitive_kind is not None,
                sensitive_kind=sensitive_kind,
                source_sheet=sheet_name,
                source_row=row_index,
                source_file=source_file,
            ),
        )
    return facts


def read_named_sheet_rows(path: Path, sheet_paths: dict[str, str], sheet_name: str) -> dict[int, dict[int, str | None]]:
    return read_xlsx_rows(path, sheet_paths[sheet_name])


def serialize_raw_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def classify_raw_value(value: Any) -> tuple[str, str | None, Decimal | None, date | None, bool | None]:
    if isinstance(value, bool):
        return "bool", None, None, None, value
    if isinstance(value, datetime):
        return "date", value.isoformat(), None, value.date(), None
    if isinstance(value, date):
        return "date", value.isoformat(), None, value, None
    number = parse_decimal(value)
    if number is not None:
        return "number", str(value), number, None, None
    return "text", normalize_text(value), None, None, None


def parse_model_raw_sheets(path: Path, project: str, snapshot_month: date) -> tuple[list[ModelRawSheet], list[ModelRawRow], list[ModelRawCell]]:
    workbook = load_workbook(path, data_only=True, read_only=True, keep_links=False)
    sheets: list[ModelRawSheet] = []
    rows_out: list[ModelRawRow] = []
    cells_out: list[ModelRawCell] = []

    for worksheet in workbook.worksheets:
        sheet_kind = RAW_MODEL_SHEETS.get(worksheet.title)
        if sheet_kind is None:
            continue

        sheet_row_count = 0
        sheet_cell_count = 0
        for row_number, row in enumerate(worksheet.iter_rows(values_only=True), 1):
            raw_values: dict[str, Any] = {}
            row_label = None
            row_kind = None
            row_cell_count = 0

            for column_number, value in enumerate(row, 1):
                text = normalize_text(value)
                if text is None:
                    continue
                if row_label is None:
                    row_label = text
                raw_values[f"column_{column_number}"] = serialize_raw_value(value)
                value_type, value_text, value_number, value_date, value_bool = classify_raw_value(value)
                sensitive_kind = row_sensitive_kind(value)
                row_kind = row_kind or sensitive_kind
                cells_out.append(
                    ModelRawCell(
                        project=project,
                        snapshot_month=snapshot_month,
                        source_file=path.name,
                        sheet_name=worksheet.title,
                        sheet_kind=sheet_kind,
                        row_number=row_number,
                        column_number=column_number,
                        column_letter=get_column_letter(column_number),
                        value_type=value_type,
                        value_text=value_text,
                        value_number=value_number,
                        value_date=value_date,
                        value_bool=value_bool,
                        is_sensitive=sensitive_kind is not None,
                        sensitive_kind=sensitive_kind,
                    ),
                )
                row_cell_count += 1
                sheet_cell_count += 1

            if row_cell_count == 0:
                continue

            rows_out.append(
                ModelRawRow(
                    project=project,
                    snapshot_month=snapshot_month,
                    source_file=path.name,
                    sheet_name=worksheet.title,
                    sheet_kind=sheet_kind,
                    row_number=row_number,
                    row_label=row_label,
                    non_empty_cells=row_cell_count,
                    raw_values=raw_values,
                    is_sensitive=row_kind is not None,
                    sensitive_kind=row_kind,
                ),
            )
            sheet_row_count += 1

        sheets.append(
            ModelRawSheet(
                project=project,
                snapshot_month=snapshot_month,
                source_file=path.name,
                sheet_name=worksheet.title,
                sheet_kind=sheet_kind,
                max_row=worksheet.max_row,
                max_column=worksheet.max_column,
                row_count=sheet_row_count,
                cell_count=sheet_cell_count,
            ),
        )

    workbook.close()
    return sheets, rows_out, cells_out


def parse_model_file(path: Path, project: str) -> tuple[ModelSource, list[Any]]:
    snapshot_month = parse_snapshot_month(path)
    sheet_paths = read_sheet_paths(path)
    source = ModelSource(
        project=project,
        snapshot_month=snapshot_month,
        file_name=path.name,
        file_hash=file_hash(path),
    )
    facts: list[Any] = []

    if SHEET_FM in sheet_paths:
        facts.extend(build_monthly_facts(read_named_sheet_rows(path, sheet_paths, SHEET_FM), SHEET_FM, project, snapshot_month, "current", path.name))
    if SHEET_FM_PLAN in sheet_paths:
        facts.extend(build_monthly_facts(read_named_sheet_rows(path, sheet_paths, SHEET_FM_PLAN), SHEET_FM_PLAN, project, snapshot_month, "plan", path.name))
    if SHEET_KPI in sheet_paths:
        facts.extend(build_kpi_facts(read_named_sheet_rows(path, sheet_paths, SHEET_KPI), SHEET_KPI, project, snapshot_month, "current", path.name))
    if SHEET_KPI_PLAN in sheet_paths:
        facts.extend(build_kpi_facts(read_named_sheet_rows(path, sheet_paths, SHEET_KPI_PLAN), SHEET_KPI_PLAN, project, snapshot_month, "plan", path.name))
    if SHEET_COMPARISON in sheet_paths:
        facts.extend(build_comparison_facts(read_named_sheet_rows(path, sheet_paths, SHEET_COMPARISON), SHEET_COMPARISON, project, snapshot_month, path.name))
    if SHEET_PASSPORT in sheet_paths:
        facts.extend(build_passport_facts(read_named_sheet_rows(path, sheet_paths, SHEET_PASSPORT), SHEET_PASSPORT, project, snapshot_month, path.name))
    if SHEET_RATES in sheet_paths:
        facts.extend(build_assumption_facts(read_named_sheet_rows(path, sheet_paths, SHEET_RATES), SHEET_RATES, project, snapshot_month, path.name))
    raw_sheets, raw_rows, raw_cells = parse_model_raw_sheets(path, project, snapshot_month)
    facts.extend(raw_sheets)
    facts.extend(raw_rows)
    facts.extend(raw_cells)
    return source, facts


def replace_model_rows(session: Session, project: str, snapshot_months: set[date], sources: list[ModelSource], facts: list[Any]) -> int:
    if snapshot_months:
        for model in (
            ModelSource,
            ModelMonthlyFact,
            ModelKpiFact,
            ModelComparisonFact,
            ModelPassportFact,
            ModelAssumptionFact,
            ModelRawSheet,
            ModelRawRow,
            ModelRawCell,
        ):
            session.execute(
                delete(model).where(
                    model.project == project,
                    model.snapshot_month.in_(snapshot_months),
                ),
            )

    session.bulk_save_objects(sources)
    for model in (
        ModelMonthlyFact,
        ModelKpiFact,
        ModelComparisonFact,
        ModelPassportFact,
        ModelAssumptionFact,
        ModelRawSheet,
        ModelRawRow,
        ModelRawCell,
    ):
        model_facts = [fact for fact in facts if isinstance(fact, model)]
        if model_facts:
            session.bulk_save_objects(model_facts)
    session.commit()
    return len(facts)


def insert_mapping_batches(session: Session, model: Any, rows: Any, batch_size: int = 5000) -> int:
    batch = []
    count = 0
    for row in rows:
        batch.append(row)
        if len(batch) >= batch_size:
            session.execute(model.__table__.insert(), batch)
            count += len(batch)
            batch.clear()
    if batch:
        session.execute(model.__table__.insert(), batch)
        count += len(batch)
    return count


def object_mapping(item: Any) -> dict[str, Any]:
    skipped = {"id", "created_at", "imported_at"}
    return {column.name: getattr(item, column.name) for column in item.__table__.columns if column.name not in skipped}


def delete_model_snapshot(session: Session, project: str, snapshot_month: date) -> None:
    for model in (
        ModelSource,
        ModelMonthlyFact,
        ModelKpiFact,
        ModelComparisonFact,
        ModelPassportFact,
        ModelAssumptionFact,
        ModelRawSheet,
        ModelRawRow,
        ModelRawCell,
    ):
        session.execute(
            delete(model).where(
                model.project == project,
                model.snapshot_month == snapshot_month,
            ),
        )


def import_model(session: Session, source: Path, project: str, monthly_mode: str = "actual") -> ModelImportResult:
    files = find_xlsx_files(source)
    sources_count = 0
    monthly_count = 0
    kpi_count = 0
    comparison_count = 0
    passport_count = 0
    assumption_count = 0
    raw_sheet_count = 0
    raw_row_count = 0
    raw_cell_count = 0
    for file_path in files:
        snapshot_month = parse_snapshot_month(file_path)
        delete_model_snapshot(session, project, snapshot_month)
        session.add(
            ModelSource(
                project=project,
                snapshot_month=snapshot_month,
                file_name=file_path.name,
                file_hash=file_hash(file_path),
            ),
        )
        sources_count += 1

        sheet_paths = read_sheet_paths(file_path)
        if monthly_mode != "none" and SHEET_FM in sheet_paths:
            monthly_count += insert_mapping_batches(
                session,
                ModelMonthlyFact,
                iter_monthly_fact_mappings(
                    read_named_sheet_rows(file_path, sheet_paths, SHEET_FM),
                    SHEET_FM,
                    project,
                    snapshot_month,
                    "current",
                    file_path.name,
                    monthly_mode,
                ),
            )
        if monthly_mode != "none" and SHEET_FM_PLAN in sheet_paths:
            monthly_count += insert_mapping_batches(
                session,
                ModelMonthlyFact,
                iter_monthly_fact_mappings(
                    read_named_sheet_rows(file_path, sheet_paths, SHEET_FM_PLAN),
                    SHEET_FM_PLAN,
                    project,
                    snapshot_month,
                    "plan",
                    file_path.name,
                    monthly_mode,
                ),
            )
        if SHEET_KPI in sheet_paths:
            kpi_count += insert_mapping_batches(
                session,
                ModelKpiFact,
                (
                    object_mapping(item)
                    for item in build_kpi_facts(
                        read_named_sheet_rows(file_path, sheet_paths, SHEET_KPI),
                        SHEET_KPI,
                        project,
                        snapshot_month,
                        "current",
                        file_path.name,
                    )
                ),
            )
        if SHEET_KPI_PLAN in sheet_paths:
            kpi_count += insert_mapping_batches(
                session,
                ModelKpiFact,
                (
                    object_mapping(item)
                    for item in build_kpi_facts(
                        read_named_sheet_rows(file_path, sheet_paths, SHEET_KPI_PLAN),
                        SHEET_KPI_PLAN,
                        project,
                        snapshot_month,
                        "plan",
                        file_path.name,
                    )
                ),
            )
        if SHEET_COMPARISON in sheet_paths:
            comparison_count += insert_mapping_batches(
                session,
                ModelComparisonFact,
                (
                    object_mapping(item)
                    for item in build_comparison_facts(
                        read_named_sheet_rows(file_path, sheet_paths, SHEET_COMPARISON),
                        SHEET_COMPARISON,
                        project,
                        snapshot_month,
                        file_path.name,
                    )
                ),
            )
        if SHEET_PASSPORT in sheet_paths:
            passport_count += insert_mapping_batches(
                session,
                ModelPassportFact,
                (
                    object_mapping(item)
                    for item in build_passport_facts(
                        read_named_sheet_rows(file_path, sheet_paths, SHEET_PASSPORT),
                        SHEET_PASSPORT,
                        project,
                        snapshot_month,
                        file_path.name,
                    )
                ),
            )
        if SHEET_RATES in sheet_paths:
            assumption_count += insert_mapping_batches(
                session,
                ModelAssumptionFact,
                (
                    object_mapping(item)
                    for item in build_assumption_facts(
                        read_named_sheet_rows(file_path, sheet_paths, SHEET_RATES),
                        SHEET_RATES,
                        project,
                        snapshot_month,
                        file_path.name,
                    )
                ),
            )
        raw_sheets, raw_rows, raw_cells = parse_model_raw_sheets(file_path, project, snapshot_month)
        raw_sheet_count += insert_mapping_batches(session, ModelRawSheet, (object_mapping(item) for item in raw_sheets))
        raw_row_count += insert_mapping_batches(session, ModelRawRow, (object_mapping(item) for item in raw_rows))
        raw_cell_count += insert_mapping_batches(session, ModelRawCell, (object_mapping(item) for item in raw_cells))
        session.commit()

    return ModelImportResult(
        files=len(files),
        sources=sources_count,
        monthly=monthly_count,
        kpi=kpi_count,
        comparison=comparison_count,
        passport=passport_count,
        assumptions=assumption_count,
        raw_sheets=raw_sheet_count,
        raw_rows=raw_row_count,
        raw_cells=raw_cell_count,
    )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--project", default=DEFAULT_PROJECT)
    parser.add_argument("--monthly-mode", choices=("actual", "full", "none"), default="actual")
    args = parser.parse_args()

    settings = load_settings()
    create_database_tables(settings.database_url)
    session_factory = get_session_factory(settings.database_url)
    with session_factory() as session:
        result = import_model(session, args.source, args.project, args.monthly_mode)

    print(
        "Imported "
        f"{result.files} files, "
        f"{result.monthly} monthly facts, "
        f"{result.kpi} KPI facts, "
        f"{result.comparison} comparison facts, "
        f"{result.passport} passport facts, "
        f"{result.assumptions} assumption facts, "
        f"{result.raw_sheets} raw sheets, "
        f"{result.raw_rows} raw rows, "
        f"{result.raw_cells} raw cells"
    )


if __name__ == "__main__":
    main()
