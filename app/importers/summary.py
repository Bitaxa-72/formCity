import argparse
import hashlib
import re
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import delete
from sqlalchemy.orm import Session

from app.core.config import load_settings
from app.db.database import create_database_tables, get_session_factory
from app.db.models import SummaryCell, SummaryRow, SummarySheet, SummarySource


DEFAULT_SOURCE = Path("..") / "оригиналы таблиц" / "сводная"
PROJECT_ALIASES = {
    "moskovsky": ("москов", "велл"),
    "obvodny": ("обвод", "118"),
    "evgenievsky": ("евгеньев",),
}
HEADER_KEYWORDS = (
    "фио",
    "клиент",
    "№",
    "номер",
    "помещ",
    "апарт",
    "кв",
    "этаж",
    "площад",
    "дата",
    "цена",
    "оплат",
    "остат",
    "дду",
    "договор",
    "корпус",
    "класс",
)
SENSITIVE_HEADER_PARTS = (
    "фио",
    "клиент",
    "агент",
    "менеджер",
    "дду",
    "договор",
    "дкп",
    "бронь",
    "примеч",
    "контакт",
)
MONTH_LABEL_RE = re.compile(
    r"^(январь|февраль|март|апрель|май|июнь|июль|август|сентябрь|октябрь|ноябрь|декабрь)\s+\d{4}$",
    re.IGNORECASE,
)


@dataclass(frozen=True)
class SummaryImportResult:
    files: int
    sources: int
    sheets: int
    rows: int
    cells: int


def normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    text = " ".join(str(value).strip().split())
    return text or None


def normalize_search_text(value: Any) -> str:
    return " ".join((str(value) if value is not None else "").strip().lower().replace("ё", "е").split())


def normalize_key(value: Any) -> str | None:
    text = normalize_search_text(value)
    if not text:
        return None
    text = re.sub(r"[^0-9a-zа-я]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text[:255] or None


def parse_decimal(value: Any) -> Decimal | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int | float):
        return Decimal(str(value))
    if isinstance(value, str):
        prepared = value.strip().replace(" ", "").replace("\xa0", "").replace(",", ".")
        if not prepared:
            return None
        try:
            return Decimal(prepared)
        except InvalidOperation:
            return None
    return None


def serialize_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def find_xlsx_files(source: Path) -> list[Path]:
    return sorted(path for path in source.rglob("*.xlsx") if not path.name.startswith("~$"))


def infer_project(path: Path, source_root: Path) -> str:
    relative_text = normalize_search_text(path.relative_to(source_root))
    for project, aliases in PROJECT_ALIASES.items():
        if any(alias in relative_text for alias in aliases):
            return project
    return "all"


def sheet_kind_for(sheet_name: str) -> str:
    normalized = normalize_search_text(sheet_name)
    if "апартамент" in normalized or "квартир" in normalized:
        return "residential_units"
    if "коммерц" in normalized:
        return "commercial_units"
    if "кладов" in normalized:
        return "storage_units"
    if "расторж" in normalized:
        return "contract_termination"
    if "уступ" in normalized:
        return "assignment"
    if "гд" in normalized or "гарант" in normalized or "аренд" in normalized:
        return "guaranteed_income"
    if "дат" in normalized or "тек" in normalized:
        return "timeline"
    if "класс" in normalized:
        return "class_summary"
    if "свод" in normalized:
        return "summary_totals"
    if "дкп" in normalized:
        return "sale_purchase_contract"
    if "окн" in normalized:
        return "window_agreements"
    if "агент" in normalized:
        return "agents"
    return "generic"


def value_at(rows: list[tuple[Any, ...]], row: int, column: int) -> Any:
    row_values = rows[row - 1]
    index = column - 1
    if index < 0 or index >= len(row_values):
        return None
    return row_values[index]


def non_empty_values(row: tuple[Any, ...]) -> list[Any]:
    return [value for value in row if normalize_text(value) is not None]


def detect_header_row(rows: list[tuple[Any, ...]]) -> int | None:
    best_row = None
    best_score = 0
    for row_number, row in enumerate(rows[:10], 1):
        values = non_empty_values(row)
        if not values:
            continue
        text = normalize_search_text(" ".join(str(value) for value in values))
        keyword_score = sum(1 for keyword in HEADER_KEYWORDS if keyword in text)
        score = len(values) * 2 + keyword_score * 10
        if score > best_score:
            best_score = score
            best_row = row_number
    return best_row


def build_headers(rows: list[tuple[Any, ...]], header_row: int | None, max_column: int) -> dict[int, tuple[str | None, str | None]]:
    headers = {}
    for column in range(1, max_column + 1):
        label = normalize_text(value_at(rows, header_row, column)) if header_row is not None else None
        headers[column] = (label, normalize_key(label))
    return headers


def classify_value(value: Any) -> tuple[str, str | None, Decimal | None, date | None, bool | None]:
    if isinstance(value, bool):
        return "bool", None, None, None, value
    if isinstance(value, datetime):
        return "date", value.isoformat(), None, value.date(), None
    if isinstance(value, date):
        return "date", value.isoformat(), None, value, None
    number = parse_decimal(value)
    if number is not None:
        return "number", str(value), number, None, None
    return "text", normalize_text(value), None, None, None


def is_sensitive_header(header_label: str | None) -> bool:
    normalized = normalize_search_text(header_label)
    return any(part in normalized for part in SENSITIVE_HEADER_PARTS)


def find_value_by_header(raw_values: dict[str, Any], *needles: str) -> str | None:
    for key, value in raw_values.items():
        normalized = normalize_search_text(key)
        if all(needle in normalized for needle in needles):
            return normalize_text(value)
    return None


def classify_row(row_number: int, header_row: int | None, row_label: str | None, non_empty_count: int) -> str:
    if header_row is not None and row_number <= header_row:
        return "header"
    if row_label and MONTH_LABEL_RE.match(row_label):
        return "period_group"
    if non_empty_count <= 2 and row_label:
        return "group"
    return "detail"


def parse_summary_file(
    path: Path,
    source_root: Path,
) -> tuple[SummarySource, list[SummarySheet], list[SummaryRow], list[SummaryCell]]:
    project = infer_project(path, source_root)
    workbook = load_workbook(path, data_only=True, read_only=True, keep_links=False)
    source = SummarySource(project=project, file_name=path.name, file_hash=file_hash(path))
    sheets = []
    rows_out = []
    cells_out = []

    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            continue
        header_row = detect_header_row(rows)
        headers = build_headers(rows, header_row, worksheet.max_column)
        sheet_kind = sheet_kind_for(worksheet.title)
        sheet_row_count = 0
        sheet_cell_count = 0

        for row_number, row in enumerate(rows, 1):
            values = non_empty_values(row)
            if not values:
                continue
            raw_values = {}
            row_sensitive = False
            row_label = None

            for column_number, value in enumerate(row, 1):
                text = normalize_text(value)
                if text is None:
                    continue
                if row_label is None:
                    row_label = text
                header_label, header_key = headers.get(column_number, (None, None))
                raw_key = header_label or f"column_{column_number}"
                raw_values[raw_key] = serialize_value(value)
                value_type, value_text, value_number, value_date, value_bool = classify_value(value)
                cell_sensitive = is_sensitive_header(header_label)
                row_sensitive = row_sensitive or cell_sensitive
                cells_out.append(
                    SummaryCell(
                        project=project,
                        source_file=path.name,
                        sheet_name=worksheet.title,
                        sheet_kind=sheet_kind,
                        row_number=row_number,
                        column_number=column_number,
                        column_letter=get_column_letter(column_number),
                        header_row=header_row,
                        header_label=header_label,
                        header_key=header_key,
                        value_type=value_type,
                        value_text=value_text,
                        value_number=value_number,
                        value_date=value_date,
                        value_bool=value_bool,
                        is_sensitive=cell_sensitive,
                    ),
                )
                sheet_cell_count += 1

            customer_name = find_value_by_header(raw_values, "фио") or find_value_by_header(raw_values, "клиент")
            unit_number = (
                find_value_by_header(raw_values, "помещ")
                or find_value_by_header(raw_values, "апарт")
                or find_value_by_header(raw_values, "кв")
                or find_value_by_header(raw_values, "номер")
            )
            period_label = row_label if row_label and MONTH_LABEL_RE.match(row_label) else None
            row_type = classify_row(row_number, header_row, row_label, len(values))
            if customer_name:
                row_sensitive = True

            rows_out.append(
                SummaryRow(
                    project=project,
                    source_file=path.name,
                    sheet_name=worksheet.title,
                    sheet_kind=sheet_kind,
                    row_number=row_number,
                    row_type=row_type,
                    row_label=row_label,
                    period_label=period_label,
                    unit_number=unit_number,
                    customer_name=customer_name,
                    non_empty_cells=len(values),
                    raw_values=raw_values,
                    is_sensitive=row_sensitive,
                    sensitive_fields={
                        "customer_name": bool(customer_name),
                        "unit_number": bool(unit_number),
                    },
                ),
            )
            sheet_row_count += 1

        sheets.append(
            SummarySheet(
                project=project,
                source_file=path.name,
                sheet_name=worksheet.title,
                sheet_kind=sheet_kind,
                header_row=header_row,
                max_row=worksheet.max_row,
                max_column=worksheet.max_column,
                row_count=sheet_row_count,
                cell_count=sheet_cell_count,
            ),
        )

    return source, sheets, rows_out, cells_out


def replace_summary(
    session: Session,
    sources: list[SummarySource],
    sheets: list[SummarySheet],
    rows: list[SummaryRow],
    cells: list[SummaryCell],
) -> tuple[int, int, int]:
    projects_and_files = {(source.project, source.file_name) for source in sources}
    for project, source_file in projects_and_files:
        for model in (SummaryCell, SummaryRow, SummarySheet, SummarySource):
            file_column = model.file_name if model is SummarySource else model.source_file
            session.execute(
                delete(model).where(
                    model.project == project,
                    file_column == source_file,
                ),
            )

    session.add_all(sources)
    session.add_all(sheets)
    session.add_all(rows)
    session.add_all(cells)
    session.commit()
    return len(sheets), len(rows), len(cells)


def import_summary(session: Session, source: Path) -> SummaryImportResult:
    files = find_xlsx_files(source)
    sources = []
    sheets = []
    rows = []
    cells = []
    for file_path in files:
        parsed_source, parsed_sheets, parsed_rows, parsed_cells = parse_summary_file(file_path, source)
        sources.append(parsed_source)
        sheets.extend(parsed_sheets)
        rows.extend(parsed_rows)
        cells.extend(parsed_cells)

    sheet_count, row_count, cell_count = replace_summary(session, sources, sheets, rows, cells)
    return SummaryImportResult(files=len(files), sources=len(sources), sheets=sheet_count, rows=row_count, cells=cell_count)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    args = parser.parse_args()

    settings = load_settings()
    create_database_tables(settings.database_url)
    session_factory = get_session_factory(settings.database_url)
    with session_factory() as session:
        result = import_summary(session, args.source)

    print(f"Imported {result.files} files, {result.sheets} summary sheets, {result.rows} rows, {result.cells} cells")


if __name__ == "__main__":
    main()
