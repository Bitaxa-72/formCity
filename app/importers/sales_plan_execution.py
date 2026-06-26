import argparse
import hashlib
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import delete
from sqlalchemy.orm import Session

from app.core.config import load_settings
from app.db.database import create_database_tables, get_session_factory
from app.db.models import SalesPlanExecutionFact, SalesPlanExecutionSource


DEFAULT_SOURCE = Path("..") / "оригиналы таблиц" / "отчет об исполнении плана продаж"
DEFAULT_PROJECT = "obvodny"
SOURCE_DATE_RE = re.compile(r"(\d{2})\.(\d{2})\.(\d{4})")
YEAR_RE = re.compile(r"\b(20\d{2})\b")

SEGMENT_BY_LABEL = {
    "итого проект": ("project_total", "Итого проект"),
    "апартаменты": ("apartments", "Апартаменты"),
    "ресторан": ("restaurant", "Ресторан"),
}

SCENARIOS_BY_BLOCK = {
    "segment_cumulative": {
        4: "plan",
        5: "fact",
        6: "deviation",
        8: "fact_forecast",
        9: "forecast_deviation",
    },
    "month": {
        4: "plan",
        5: "forecast",
        6: "fact",
        7: "fact_minus_forecast",
    },
    "year": {
        4: "plan",
        5: "forecast",
        6: "fact_actualized_forecast",
        7: "fact",
        8: "remaining_to_sell",
    },
    "project_lifetime": {
        4: "plan",
        5: "forecast",
        6: "fact_actualized_forecast",
        7: "fact",
        8: "remaining_to_sell",
    },
}


@dataclass(frozen=True)
class SalesPlanExecutionImportResult:
    files: int
    sources: int
    facts: int


@dataclass(frozen=True)
class Block:
    start_row: int
    end_row: int
    block_kind: str
    block_label: str
    segment: str
    segment_label: str
    period_kind: str
    period_month: date | None
    year: int | None


def normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    text = " ".join(str(value).strip().split())
    return text or None


def normalize_search_text(value: str | None) -> str:
    return " ".join((value or "").strip().lower().replace("ё", "е").split())


def parse_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int | float):
        return Decimal(str(value))
    if isinstance(value, str):
        prepared = value.strip().replace(" ", "").replace(",", ".")
        if not prepared:
            return None
        try:
            return Decimal(prepared)
        except InvalidOperation:
            return None
    return None


def parse_date_value(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value[:19]).date()
        except ValueError:
            return None
    return None


def month_start(value: date) -> date:
    return date(value.year, value.month, 1)


def parse_snapshot_month_from_filename(path: Path) -> date | None:
    match = SOURCE_DATE_RE.search(path.name)
    if match is None:
        return None
    return date(int(match.group(3)), int(match.group(2)), 1)


def file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def metric_key_for(label: str) -> tuple[str, str, str] | None:
    normalized = normalize_search_text(label)
    if normalized == "продажи, руб.":
        return "sales_revenue", "Продажи", "rub"
    if normalized == "поступление ден. средств, руб.":
        return "cash_receipts", "Поступление денежных средств", "rub"
    if normalized == "объем законтрактованных площадей, м2":
        return "contract_area_sqm", "Объем законтрактованных площадей", "sqm"
    if normalized == "объем законтрактованных площадей, шт":
        return "contract_count", "Объем законтрактованных площадей", "count"
    if normalized == "цена за 1 м2, руб.":
        return "price_per_sqm", "Цена за 1 м2", "rub_per_sqm"
    return None


def owner_scope_for(label: str) -> str:
    normalized = normalize_search_text(label)
    if normalized == "в т.ч. на застройщика, руб.":
        return "developer"
    if normalized == "в т.ч. на велл (уступка), руб.":
        return "well"
    return "all"


def find_xlsx_files(source: Path) -> list[Path]:
    return sorted(path for path in source.rglob("*.xlsx") if not path.name.startswith("~$"))


def extract_snapshot_date(rows: list[tuple[Any, ...]]) -> date | None:
    for row in rows[:5]:
        for value in row:
            if not isinstance(value, str):
                continue
            match = SOURCE_DATE_RE.search(value)
            if match:
                return date(int(match.group(3)), int(match.group(2)), int(match.group(1)))
    return None


def find_work_sheet(workbook) -> Any:
    for sheet in workbook.worksheets:
        if sheet.max_row > 1 and sheet.max_column > 1:
            return sheet
    return workbook.active


def build_blocks(rows: list[tuple[Any, ...]], snapshot_month: date) -> list[Block]:
    def value_at(row: int, column: int) -> Any:
        row_values = rows[row - 1]
        index = column - 1
        if index >= len(row_values):
            return None
        return row_values[index]

    headers: list[tuple[int, str, str, str, str, str, date | None, int | None]] = []
    for row in range(1, len(rows) + 1):
        label = normalize_text(value_at(row, 3))
        normalized = normalize_search_text(label)
        column_4 = normalize_search_text(normalize_text(value_at(row, 4)))
        column_5 = normalize_search_text(normalize_text(value_at(row, 5)))
        column_6 = normalize_search_text(normalize_text(value_at(row, 6)))
        column_7 = normalize_search_text(normalize_text(value_at(row, 7)))
        column_8 = normalize_search_text(normalize_text(value_at(row, 8)))
        column_9 = normalize_search_text(normalize_text(value_at(row, 9)))

        if normalized in SEGMENT_BY_LABEL and column_4 == "план" and column_5 == "факт" and column_6 == "отклонение" and column_8 == "факт+прогноз":
            segment, segment_label = SEGMENT_BY_LABEL[normalized]
            headers.append((row, "segment_cumulative", label or segment_label, segment, segment_label, "snapshot", snapshot_month, None))
            continue

        period_date = parse_date_value(value_at(row, 3))
        if period_date and column_4 == "план" and column_5 == "прогноз" and column_6 == "факт":
            headers.append((row, "month", "Месяц", "project_total", "Итого проект", "month", month_start(period_date), None))
            continue

        if normalized == "итого проект" and column_4 == "план" and column_5 == "прогноз" and column_7 == "факт" and column_8 == "остаток к продаже":
            headers.append((row, "project_lifetime", label or "Итого проект", "project_total", "Итого проект", "project_total", None, None))
            continue

        if normalized and normalized.startswith("итого ") and column_4 == "план" and column_5 == "прогноз" and column_7 == "факт":
            match = YEAR_RE.search(normalized)
            year = int(match.group(1)) if match else None
            headers.append((row, "year", label or "Итого за год", "project_total", "Итого проект", "year", None, year))
            continue

    blocks = []
    for index, header in enumerate(headers):
        next_row = headers[index + 1][0] if index + 1 < len(headers) else len(rows) + 1
        row, block_kind, block_label, segment, segment_label, period_kind, period_month, year = header
        blocks.append(
            Block(
                start_row=row + 1,
                end_row=next_row - 1,
                block_kind=block_kind,
                block_label=block_label,
                segment=segment,
                segment_label=segment_label,
                period_kind=period_kind,
                period_month=period_month,
                year=year,
            ),
        )
    return blocks


def parse_sales_plan_execution_file(path: Path, project: str) -> tuple[SalesPlanExecutionSource, list[SalesPlanExecutionFact]]:
    workbook = load_workbook(path, data_only=True, read_only=True, keep_links=False)
    sheet = find_work_sheet(workbook)
    rows = list(sheet.iter_rows(values_only=True))

    def value_at(row: int, column: int) -> Any:
        row_values = rows[row - 1]
        index = column - 1
        if index >= len(row_values):
            return None
        return row_values[index]

    snapshot_date = extract_snapshot_date(rows)
    snapshot_month = parse_snapshot_month_from_filename(path) or (month_start(snapshot_date) if snapshot_date else None)
    if snapshot_month is None:
        raise ValueError("sales_plan_execution_period_not_found")

    source = SalesPlanExecutionSource(
        project=project,
        snapshot_month=snapshot_month,
        snapshot_date=snapshot_date,
        file_name=path.name,
        file_hash=file_hash(path),
    )

    facts = []
    for block in build_blocks(rows, snapshot_month):
        current_metric: tuple[str, str, str] | None = None
        for row in range(block.start_row, block.end_row + 1):
            label = normalize_text(value_at(row, 3))
            if label is None:
                continue

            metric = metric_key_for(label)
            if metric is not None:
                current_metric = metric
                owner_scope = "all"
            elif current_metric is not None:
                owner_scope = owner_scope_for(label)
                if owner_scope == "all":
                    continue
                metric = current_metric
            else:
                continue

            metric_key, metric_name, unit = metric
            for column, scenario in SCENARIOS_BY_BLOCK[block.block_kind].items():
                value = parse_decimal(value_at(row, column))
                if value is None:
                    continue
                facts.append(
                    SalesPlanExecutionFact(
                        project=project,
                        snapshot_month=snapshot_month,
                        snapshot_date=snapshot_date,
                        block_kind=block.block_kind,
                        block_label=block.block_label,
                        segment=block.segment,
                        segment_label=block.segment_label,
                        metric_key=metric_key,
                        metric_name=metric_name,
                        owner_scope=owner_scope,
                        period_kind=block.period_kind,
                        period_month=block.period_month,
                        year=block.year,
                        scenario=scenario,
                        value=value,
                        unit=unit,
                        source_sheet=sheet.title,
                        source_row=row,
                        source_col=column,
                        source_file=path.name,
                    ),
                )

    return source, facts


def replace_sales_plan_execution(
    session: Session,
    project: str,
    sources: list[SalesPlanExecutionSource],
    facts: list[SalesPlanExecutionFact],
) -> int:
    periods = {source.snapshot_month for source in sources}
    if periods:
        session.execute(
            delete(SalesPlanExecutionFact).where(
                SalesPlanExecutionFact.project == project,
                SalesPlanExecutionFact.snapshot_month.in_(periods),
            ),
        )
        session.execute(
            delete(SalesPlanExecutionSource).where(
                SalesPlanExecutionSource.project == project,
                SalesPlanExecutionSource.snapshot_month.in_(periods),
            ),
        )

    session.add_all(sources)
    session.add_all(facts)
    session.commit()
    return len(facts)


def import_sales_plan_execution(session: Session, source: Path, project: str) -> SalesPlanExecutionImportResult:
    files = find_xlsx_files(source)
    sources = []
    facts = []
    for file_path in files:
        parsed_source, parsed_facts = parse_sales_plan_execution_file(file_path, project)
        sources.append(parsed_source)
        facts.extend(parsed_facts)

    fact_count = replace_sales_plan_execution(session, project, sources, facts)
    return SalesPlanExecutionImportResult(files=len(files), sources=len(sources), facts=fact_count)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--project", default=DEFAULT_PROJECT)
    args = parser.parse_args()

    settings = load_settings()
    create_database_tables(settings.database_url)
    session_factory = get_session_factory(settings.database_url)
    with session_factory() as session:
        result = import_sales_plan_execution(session, args.source, args.project)

    print(f"Imported {result.files} files, {result.facts} sales plan execution facts")


if __name__ == "__main__":
    main()
