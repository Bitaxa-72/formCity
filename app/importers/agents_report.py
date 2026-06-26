import argparse
import hashlib
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import delete
from sqlalchemy.orm import Session

from app.core.config import load_settings
from app.db.database import create_database_tables, get_session_factory
from app.db.models import AgentsReportDeal, AgentsReportMonthlyValue, AgentsReportSource


DEFAULT_SOURCE = Path("..") / "оригиналы таблиц" / "отчет по агентам"
DEFAULT_PROJECT = "obvodny"
SOURCE_DATE_RE = re.compile(r"(\d{2})\.(\d{2})\.(\d{4})")


@dataclass(frozen=True)
class AgentsReportImportResult:
    files: int
    sources: int
    deals: int
    monthly_values: int


def normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    text = " ".join(str(value).strip().split())
    return text or None


def normalize_search_text(value: Any) -> str:
    return " ".join((str(value) if value is not None else "").strip().lower().replace("ё", "е").split())


def parse_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int | float):
        return Decimal(str(value))
    if isinstance(value, str):
        prepared = value.strip().replace(" ", "").replace(",", ".")
        if not prepared:
            return None
        try:
            return Decimal(prepared)
        except InvalidOperation:
            return None
    return None


def parse_date_value(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        stripped = value.strip()
        match = SOURCE_DATE_RE.search(stripped)
        if match:
            return date(int(match.group(3)), int(match.group(2)), int(match.group(1)))
        try:
            return datetime.fromisoformat(stripped[:19]).date()
        except ValueError:
            return None
    return None


def month_start(value: date) -> date:
    return date(value.year, value.month, 1)


def parse_snapshot_date_from_filename(path: Path) -> date | None:
    match = SOURCE_DATE_RE.search(path.name)
    if match is None:
        return None
    return date(int(match.group(3)), int(match.group(2)), int(match.group(1)))


def file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024, ), b""):
            digest.update(chunk)
    return digest.hexdigest()


def find_xlsx_files(source: Path) -> list[Path]:
    return sorted(path for path in source.rglob("*.xlsx") if not path.name.startswith("~$"))


def find_column(headers: tuple[Any, ...], *needles: str) -> int | None:
    for index, value in enumerate(headers, 1):
        normalized = normalize_search_text(value)
        if normalized and all(needle in normalized for needle in needles):
            return index
    return None


def find_column_by(headers: tuple[Any, ...], predicate) -> int | None:
    for index, value in enumerate(headers, 1):
        normalized = normalize_search_text(value)
        if normalized and predicate(normalized):
            return index
    return None


def value_at(rows: list[tuple[Any, ...]], row: int, column: int | None) -> Any:
    if column is None:
        return None
    row_values = rows[row - 1]
    index = column - 1
    if index < 0 or index >= len(row_values):
        return None
    return row_values[index]


def build_columns(headers: tuple[Any, ...]) -> dict[str, int | None]:
    return {
        "agent_name": find_column(headers, "наименование агента"),
        "unit_number": find_column(headers, "помещ"),
        "buyer_name": find_column(headers, "фио"),
        "ddu_number": find_column(headers, "дду"),
        "contract_date": find_column(headers, "дата"),
        "area_sqm": find_column(headers, "площадь"),
        "commission_base_amount": find_column(headers, "цена сделки"),
        "check_qw_amount": find_column(headers, "проверка", "q-w"),
        "check_gh_amount": find_column(headers, "проверка", "g-h"),
        "commission_rate": find_column(headers, "размер", "вознаграждения", "%"),
        "commission_amount": find_column(headers, "размер", "вознаграждения", "руб."),
        "act_total_amount": find_column(headers, "итого сумма по акту"),
        "paid_amount": find_column(headers, "оплачено"),
        "remaining_amount": find_column(headers, "остаток к оплате"),
        "act_info": find_column(headers, "и дата", "акта"),
        "budget_month": find_column(headers, "бюджет месяц"),
        "ddu_assignment_amount": find_column_by(headers, lambda value: "дду+уступка" in value and "руб." in value and "м2" not in value),
        "ddu_assignment_price_per_sqm": find_column_by(headers, lambda value: "дду+уступка" in value and "м2" in value),
        "ddu_amount": find_column_by(headers, lambda value: value.startswith("дду") and "дду+уступка" not in value and "руб." in value and "м2" not in value),
        "ddu_price_per_sqm": find_column_by(headers, lambda value: value.startswith("дду") and "дду+уступка" not in value and "м2" in value),
        "assignment_amount": find_column_by(headers, lambda value: value.startswith("уступка") and "руб." in value and "м2" not in value),
        "assignment_price_per_sqm": find_column_by(headers, lambda value: value.startswith("уступка") and "м2" in value),
        "furniture_amount": find_column(headers, "меблировка"),
        "note": find_column(headers, "примечание"),
    }


def find_monthly_blocks(headers: tuple[Any, ...]) -> list[tuple[str, int, list[tuple[int, str, date | None]]]]:
    blocks = []
    label_columns = [index for index, value in enumerate(headers, 1) if normalize_search_text(value) == "прошлые периоды"]
    for block_index, label_column in enumerate(label_columns):
        value_kind = "ddu_schedule" if block_index == 0 else "assignment_schedule"
        columns = [(label_column, "past_periods_total", None)]
        column = label_column + 1
        while column <= len(headers):
            period = parse_date_value(headers[column - 1])
            if period is None:
                break
            columns.append((column, "month", month_start(period)))
            column += 1
        blocks.append((value_kind, label_column, columns))
    return blocks


def parse_agents_report_file(path: Path, project: str) -> tuple[AgentsReportSource, list[AgentsReportDeal], list[AgentsReportMonthlyValue]]:
    workbook = load_workbook(path, data_only=True, read_only=True, keep_links=False)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    headers = rows[2]
    columns = build_columns(headers)
    monthly_blocks = find_monthly_blocks(headers)

    snapshot_date = parse_snapshot_date_from_filename(path)
    if snapshot_date is None:
        raise ValueError("agents_report_period_not_found")
    snapshot_month = month_start(snapshot_date)

    source = AgentsReportSource(
        project=project,
        snapshot_month=snapshot_month,
        snapshot_date=snapshot_date,
        file_name=path.name,
        file_hash=file_hash(path),
    )

    deals = []
    monthly_values = []
    for row in range(4, len(rows) + 1):
        agent_name = normalize_text(value_at(rows, row, columns["agent_name"]))
        normalized_agent = normalize_search_text(agent_name)
        if agent_name is None or normalized_agent.startswith("итого"):
            continue

        unit_number = normalize_text(value_at(rows, row, columns["unit_number"]))
        buyer_name = normalize_text(value_at(rows, row, columns["buyer_name"]))
        ddu_number = normalize_text(value_at(rows, row, columns["ddu_number"]))
        act_info = normalize_text(value_at(rows, row, columns["act_info"]))
        note = normalize_text(value_at(rows, row, columns["note"]))
        if not any((unit_number, buyer_name, ddu_number, parse_decimal(value_at(rows, row, columns["commission_amount"])), parse_decimal(value_at(rows, row, columns["act_total_amount"])))):
            continue

        deal = AgentsReportDeal(
            project=project,
            snapshot_month=snapshot_month,
            snapshot_date=snapshot_date,
            row_order=row,
            agent_name=agent_name,
            unit_number=unit_number,
            buyer_name=buyer_name,
            ddu_number=ddu_number,
            contract_date=parse_date_value(value_at(rows, row, columns["contract_date"])),
            area_sqm=parse_decimal(value_at(rows, row, columns["area_sqm"])),
            commission_base_amount=parse_decimal(value_at(rows, row, columns["commission_base_amount"])),
            check_qw_amount=parse_decimal(value_at(rows, row, columns["check_qw_amount"])),
            check_gh_amount=parse_decimal(value_at(rows, row, columns["check_gh_amount"])),
            commission_rate=parse_decimal(value_at(rows, row, columns["commission_rate"])),
            commission_amount=parse_decimal(value_at(rows, row, columns["commission_amount"])),
            act_total_amount=parse_decimal(value_at(rows, row, columns["act_total_amount"])),
            paid_amount=parse_decimal(value_at(rows, row, columns["paid_amount"])),
            remaining_amount=parse_decimal(value_at(rows, row, columns["remaining_amount"])),
            act_info=act_info,
            budget_month=parse_date_value(value_at(rows, row, columns["budget_month"])),
            ddu_assignment_amount=parse_decimal(value_at(rows, row, columns["ddu_assignment_amount"])),
            ddu_assignment_price_per_sqm=parse_decimal(value_at(rows, row, columns["ddu_assignment_price_per_sqm"])),
            ddu_amount=parse_decimal(value_at(rows, row, columns["ddu_amount"])),
            ddu_price_per_sqm=parse_decimal(value_at(rows, row, columns["ddu_price_per_sqm"])),
            assignment_amount=parse_decimal(value_at(rows, row, columns["assignment_amount"])),
            assignment_price_per_sqm=parse_decimal(value_at(rows, row, columns["assignment_price_per_sqm"])),
            furniture_amount=parse_decimal(value_at(rows, row, columns["furniture_amount"])),
            note=note,
            unit="rub",
            is_sensitive=True,
            sensitive_fields={
                "agent_name": True,
                "buyer_name": True,
                "ddu_number": True,
                "act_info": True,
            },
            source_sheet=sheet.title,
            source_row=row,
            source_file=path.name,
        )
        deals.append(deal)

        for value_kind, _, month_columns in monthly_blocks:
            for column, period_kind, period_month in month_columns:
                value = parse_decimal(value_at(rows, row, column))
                if value is None:
                    continue
                monthly_values.append(
                    AgentsReportMonthlyValue(
                        project=project,
                        snapshot_month=snapshot_month,
                        snapshot_date=snapshot_date,
                        deal_source_row=row,
                        value_kind=value_kind,
                        period_kind=period_kind,
                        period_month=period_month,
                        value=value,
                        unit="rub",
                        source_sheet=sheet.title,
                        source_row=row,
                        source_col=column,
                        source_file=path.name,
                    ),
                )

    return source, deals, monthly_values


def replace_agents_report(
    session: Session,
    project: str,
    sources: list[AgentsReportSource],
    deals: list[AgentsReportDeal],
    monthly_values: list[AgentsReportMonthlyValue],
) -> tuple[int, int]:
    dates = {source.snapshot_date for source in sources}
    if dates:
        session.execute(
            delete(AgentsReportMonthlyValue).where(
                AgentsReportMonthlyValue.project == project,
                AgentsReportMonthlyValue.snapshot_date.in_(dates),
            ),
        )
        session.execute(
            delete(AgentsReportDeal).where(
                AgentsReportDeal.project == project,
                AgentsReportDeal.snapshot_date.in_(dates),
            ),
        )
        session.execute(
            delete(AgentsReportSource).where(
                AgentsReportSource.project == project,
                AgentsReportSource.snapshot_date.in_(dates),
            ),
        )

    session.add_all(sources)
    session.add_all(deals)
    session.add_all(monthly_values)
    session.commit()
    return len(deals), len(monthly_values)


def import_agents_report(session: Session, source: Path, project: str) -> AgentsReportImportResult:
    files = find_xlsx_files(source)
    sources = []
    deals = []
    monthly_values = []
    for file_path in files:
        parsed_source, parsed_deals, parsed_monthly_values = parse_agents_report_file(file_path, project)
        sources.append(parsed_source)
        deals.extend(parsed_deals)
        monthly_values.extend(parsed_monthly_values)

    deal_count, monthly_value_count = replace_agents_report(session, project, sources, deals, monthly_values)
    return AgentsReportImportResult(files=len(files), sources=len(sources), deals=deal_count, monthly_values=monthly_value_count)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--project", default=DEFAULT_PROJECT)
    args = parser.parse_args()

    settings = load_settings()
    create_database_tables(settings.database_url)
    session_factory = get_session_factory(settings.database_url)
    with session_factory() as session:
        result = import_agents_report(session, args.source, args.project)

    print(f"Imported {result.files} files, {result.deals} agent deals, {result.monthly_values} monthly values")


if __name__ == "__main__":
    main()
