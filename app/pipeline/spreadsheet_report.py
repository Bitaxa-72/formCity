from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.pipeline.calculation_engine import CalculationResult
from app.pipeline.response_data import ResponseData
from app.pipeline.sensitive_data import visible_rows


XLSX_REPORT_NOTICE = "Отчет слишком большой, оформлю вам XLSX."
XLSX_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
XLSX_REPORT_THRESHOLD = 30


def should_send_xlsx_report(response_data: ResponseData | None, calculation_result: CalculationResult | None) -> bool:
    if response_data is None or calculation_result is None:
        return False
    source = response_data.source
    return (
        calculation_result.kind == "sql_result"
        and source.get("report_type") == "model"
        and source.get("view") in {"model_raw_rows", "model_raw_search"}
        and calculation_result.row_count > XLSX_REPORT_THRESHOLD
    )


def parse_values_preview(value: object) -> dict[str, str]:
    if not isinstance(value, str):
        return {}
    parsed = {}
    for item in value.split(" | "):
        if ": " not in item:
            continue
        column, cell_value = item.split(": ", 1)
        column = column.strip()
        if column:
            parsed[column] = cell_value.strip()
    return parsed


def excel_column_index(column: str) -> int:
    result = 0
    for char in column.upper():
        if not ("A" <= char <= "Z"):
            return 10_000
        result = result * 26 + ord(char) - ord("A") + 1
    return result


def write_header(worksheet, headers: list[str]) -> None:
    fill = PatternFill("solid", fgColor="D9D9D9")
    for index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=index, value=header)
        cell.font = Font(bold=True)
        cell.fill = fill
    worksheet.freeze_panes = "A2"


def autosize_columns(worksheet, max_width: int = 60) -> None:
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        width = 10
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            width = max(width, min(len(str(value)) + 2, max_width))
        worksheet.column_dimensions[column_letter].width = width


def build_model_raw_xlsx_report(calculation_result: CalculationResult) -> tuple[bytes, str]:
    rows = visible_rows(calculation_result.rows)
    parsed_rows = [(row, parse_values_preview(row.get("values_preview"))) for row in rows]
    value_columns = sorted(
        {column for _, values in parsed_rows for column in values},
        key=excel_column_index,
    )
    headers = ["Лист", "Строка", "Название", "Открытые ячейки", *value_columns]

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "model_raw"
    write_header(worksheet, headers)

    for row_index, (row, values) in enumerate(parsed_rows, start=2):
        worksheet.cell(row=row_index, column=1, value=row.get("raw_sheet"))
        worksheet.cell(row=row_index, column=2, value=row.get("row_number"))
        worksheet.cell(row=row_index, column=3, value=row.get("row_label"))
        worksheet.cell(row=row_index, column=4, value=row.get("visible_cells"))
        for column_index, value_column in enumerate(value_columns, start=5):
            worksheet.cell(row=row_index, column=column_index, value=values.get(value_column))

    autosize_columns(worksheet)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue(), "model.xlsx"
