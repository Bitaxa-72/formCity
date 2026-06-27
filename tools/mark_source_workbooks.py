from __future__ import annotations

import shutil
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parents[2]
REPO = ROOT / "repo"
SOURCE_ROOT = ROOT / "оригиналы таблиц"
OUTPUT_ROOT = ROOT / "разметка_оригиналов"

sys.path.insert(0, str(REPO))

from app.importers import agents_report, debt_bookings, model, non_project_expenses, payment_calendar, roadmap, sales_plan_execution, sales_report, stock_for_sale, summary
from app.pipeline.sensitive_data import detect_sensitive_kind

FILL_GREEN = PatternFill("solid", fgColor="C6EFCE")
FILL_RED = PatternFill("solid", fgColor="FFC7CE")
FILL_GRAY = PatternFill("solid", fgColor="D9D9D9")
FILL_BLUE = PatternFill("solid", fgColor="BDD7EE")
FILL_WHITE = PatternFill("solid", fgColor="FFFFFF")

SAFE_MODEL_KEYS = set(model.SAFE_MODEL_METRIC_KEYS)
GREEN_MODEL_KEYS = {
    "model_revenue",
    "model_cost_of_sales",
    "model_gross_profit",
    "model_net_profit",
    "model_npv",
    "model_roe",
    "model_llcr",
    "model_total_area",
    "model_units_count",
    "model_pir",
}


def value_present(value: Any) -> bool:
    return value is not None and str(value).strip() != ""


def set_fill(cell, fill) -> None:
    cell.fill = fill


def mark_cell(ws: Worksheet, row: int | None, column: int | None, fill) -> None:
    if row and column and row > 0 and column > 0:
        set_fill(ws.cell(row=row, column=column), fill)


def mark_row_columns(ws: Worksheet, row: int, columns: list[int], fill) -> None:
    for column in columns:
        mark_cell(ws, row, column, fill)


def mark_non_empty_gray(wb) -> None:
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if value_present(cell.value):
                    set_fill(cell, FILL_GRAY)


def add_legend(wb) -> None:
    if "Разметка" in wb.sheetnames:
        del wb["Разметка"]
    ws = wb.create_sheet("Разметка", 0)
    rows = [
        ("Зеленый", "ячейка участвует в пользовательской выдаче"),
        ("Красный", "ячейка есть в БД, но не выводится по безопасности"),
        ("Серый", "ячейка не попадает в БД и не выводится"),
        ("Синий", "ячейка есть в БД, но сейчас не используется в основном публичном ответе"),
    ]
    fills = [FILL_GREEN, FILL_RED, FILL_GRAY, FILL_BLUE]
    for index, (label, description) in enumerate(rows, 1):
        ws.cell(index, 1).value = label
        ws.cell(index, 2).value = description
        ws.cell(index, 1).fill = fills[index - 1]
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 80


def first_data_sheet(wb) -> Worksheet:
    return next(ws for ws in wb.worksheets if ws.title != "Разметка")


def mark_payment_calendar(path: Path, wb) -> None:
    ws = first_data_sheet(wb)
    rows = payment_calendar.read_xlsx_rows(path)
    layout = payment_calendar.detect_payment_calendar_layout(rows)
    mark_cell(ws, 2, layout.period_column, FILL_GREEN)
    for row_index, row in rows.items():
        if row_index < 4:
            continue
        if not value_present(row.get(layout.article_column)):
            continue
        mark_row_columns(ws, row_index, [layout.article_column, layout.plan_column, layout.fact_column, layout.deviation_column], FILL_GREEN)


def mark_roadmap(path: Path, wb) -> None:
    if len(wb.worksheets) < 2:
        return
    ws = wb.worksheets[1]
    for item in roadmap.parse_roadmap_file(path, "all"):
        mark_row_columns(ws, item.row_order, [2, 3, 4, 5], FILL_GREEN)


def mark_model(path: Path, wb) -> None:
    sheet_paths = model.read_sheet_paths(path)
    rows_by_sheet = {name: model.read_named_sheet_rows(path, sheet_paths, name) for name in sheet_paths}

    def sheet(name: str) -> Worksheet | None:
        return wb[name] if name in wb.sheetnames else None

    def row_fill(sensitive_kind: str | None, metric_key: str | None) -> Any:
        if sensitive_kind:
            return FILL_RED
        if metric_key in GREEN_MODEL_KEYS:
            return FILL_GREEN
        return FILL_BLUE

    snapshot_month = model.parse_snapshot_month(path)
    if model.SHEET_FM in rows_by_sheet and sheet(model.SHEET_FM):
        ws = sheet(model.SHEET_FM)
        for row in model.iter_monthly_fact_mappings(rows_by_sheet[model.SHEET_FM], model.SHEET_FM, "obvodny", snapshot_month, "current", path.name):
            fill = row_fill(row["sensitive_kind"], row["metric_key"])
            mark_cell(ws, row["source_row"], row["source_col"], fill)
            mark_row_columns(ws, row["source_row"], [2, 3], fill)
    if model.SHEET_FM_PLAN in rows_by_sheet and sheet(model.SHEET_FM_PLAN):
        ws = sheet(model.SHEET_FM_PLAN)
        for row in model.iter_monthly_fact_mappings(rows_by_sheet[model.SHEET_FM_PLAN], model.SHEET_FM_PLAN, "obvodny", snapshot_month, "plan", path.name):
            fill = row_fill(row["sensitive_kind"], row["metric_key"])
            mark_cell(ws, row["source_row"], row["source_col"], fill)
            mark_row_columns(ws, row["source_row"], [2, 3], fill)
    for sheet_name, scenario in ((model.SHEET_KPI, "current"), (model.SHEET_KPI_PLAN, "plan")):
        if sheet_name not in rows_by_sheet or not sheet(sheet_name):
            continue
        ws = sheet(sheet_name)
        for item in model.build_kpi_facts(rows_by_sheet[sheet_name], sheet_name, "obvodny", snapshot_month, scenario, path.name):
            fill = row_fill(item.sensitive_kind, item.metric_key)
            mark_row_columns(ws, item.source_row, [5, 6], fill)
    if model.SHEET_COMPARISON in rows_by_sheet and sheet(model.SHEET_COMPARISON):
        ws = sheet(model.SHEET_COMPARISON)
        for item in model.build_comparison_facts(rows_by_sheet[model.SHEET_COMPARISON], model.SHEET_COMPARISON, "obvodny", snapshot_month, path.name):
            fill = row_fill(item.sensitive_kind, item.metric_key)
            for column in range(1, min(ws.max_column, 12) + 1):
                if value_present(ws.cell(item.source_row, column).value):
                    mark_cell(ws, item.source_row, column, fill)
    if model.SHEET_PASSPORT in rows_by_sheet and sheet(model.SHEET_PASSPORT):
        ws = sheet(model.SHEET_PASSPORT)
        for item in model.build_passport_facts(rows_by_sheet[model.SHEET_PASSPORT], model.SHEET_PASSPORT, "obvodny", snapshot_month, path.name):
            fill = row_fill(item.sensitive_kind, item.metric_key)
            mark_row_columns(ws, item.source_row, [1, 2], fill)
    if model.SHEET_RATES in rows_by_sheet and sheet(model.SHEET_RATES):
        ws = sheet(model.SHEET_RATES)
        for item in model.build_assumption_facts(rows_by_sheet[model.SHEET_RATES], model.SHEET_RATES, "obvodny", snapshot_month, path.name):
            fill = row_fill(item.sensitive_kind, item.metric_key)
            for column in range(1, ws.max_column + 1):
                if value_present(ws.cell(item.source_row, column).value):
                    mark_cell(ws, item.source_row, column, fill)
    for sheet_name in model.RAW_MODEL_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if not value_present(cell.value):
                    continue
                fill = FILL_RED if detect_sensitive_kind(str(cell.value)) else FILL_BLUE
                set_fill(cell, fill)


def mark_non_project_expenses(path: Path, wb) -> None:
    ws = first_data_sheet(wb)
    _, facts = non_project_expenses.parse_non_project_file(path, "all")
    for item in facts:
        for column in range(2, 8):
            cell = ws.cell(item.source_row, column)
            if not value_present(cell.value):
                continue
            fill = FILL_RED if detect_sensitive_kind(str(cell.value)) else FILL_GREEN
            set_fill(cell, fill)
        mark_cell(ws, 1, 3, FILL_GREEN)


def mark_stock_for_sale(path: Path, wb) -> None:
    ws = first_data_sheet(wb)
    _, facts = stock_for_sale.parse_stock_for_sale_file(path, "obvodny")
    mark_cell(ws, 4, 2, FILL_GREEN)
    for item in facts:
        mark_row_columns(ws, item.source_row, list(range(2, 9)), FILL_GREEN)


def mark_sales_report(path: Path, wb) -> None:
    ws = first_data_sheet(wb)
    _, facts = sales_report.parse_sales_report_file(path, "obvodny")
    for item in facts:
        mark_cell(ws, item.source_row, item.source_col, FILL_GREEN)
        mark_cell(ws, item.source_row, 2, FILL_GREEN)
    for row in range(1, ws.max_row + 1):
        label = sales_report.normalize_search_text(ws.cell(row, 2).value)
        if label in sales_report.SEGMENT_BY_LABEL:
            mark_cell(ws, row, 2, FILL_GREEN)


def mark_sales_plan_execution(path: Path, wb) -> None:
    parsed_wb = sales_plan_execution.load_workbook(path, data_only=True, read_only=True, keep_links=False)
    parsed_sheet = sales_plan_execution.find_work_sheet(parsed_wb)
    ws = wb[parsed_sheet.title]
    rows = list(parsed_sheet.iter_rows(values_only=True))
    snapshot_month = sales_plan_execution.parse_snapshot_month_from_filename(path)
    if snapshot_month is None:
        snapshot_date = sales_plan_execution.extract_snapshot_date(rows)
        snapshot_month = sales_plan_execution.month_start(snapshot_date) if snapshot_date else None
    if snapshot_month is not None:
        for block in sales_plan_execution.build_blocks(rows, snapshot_month):
            mark_row_columns(ws, block.start_row - 1, list(range(3, 10)), FILL_GREEN)
    _, facts = sales_plan_execution.parse_sales_plan_execution_file(path, "obvodny")
    for item in facts:
        mark_cell(ws, item.source_row, item.source_col, FILL_GREEN)
        mark_cell(ws, item.source_row, 3, FILL_GREEN)


def mark_agents_report(path: Path, wb) -> None:
    parsed_wb = agents_report.load_workbook(path, data_only=True, read_only=True, keep_links=False)
    parsed_sheet = parsed_wb.active
    ws = wb[parsed_sheet.title]
    rows = list(parsed_sheet.iter_rows(values_only=True))
    headers = rows[2]
    columns = agents_report.build_columns(headers)
    monthly_blocks = agents_report.find_monthly_blocks(headers)
    _, deals, monthly_values = agents_report.parse_agents_report_file(path, "obvodny")
    red_fields = {"agent_name", "buyer_name", "ddu_number", "act_info"}
    blue_fields = {"unit_number", "note"}
    for deal in deals:
        for field, column in columns.items():
            if column is None:
                continue
            if field in red_fields:
                fill = FILL_RED
            elif field in blue_fields:
                fill = FILL_BLUE
            else:
                fill = FILL_GREEN
            mark_cell(ws, deal.source_row, column, fill)
    for item in monthly_values:
        mark_cell(ws, item.source_row, item.source_col, FILL_GREEN)
    for _, label_column, month_columns in monthly_blocks:
        mark_cell(ws, 3, label_column, FILL_GREEN)
        for column, _, _ in month_columns:
            mark_cell(ws, 3, column, FILL_GREEN)


def mark_debt_bookings(path: Path, wb) -> None:
    parsed_wb = debt_bookings.load_workbook(path, data_only=True, read_only=True, keep_links=False)
    main_sheet = debt_bookings.find_main_sheet(parsed_wb)
    main_ws = wb[main_sheet.title]
    main_rows = list(main_sheet.iter_rows(values_only=True))
    snapshot_date = debt_bookings.parse_snapshot_date_from_sheet(main_rows) or debt_bookings.parse_snapshot_date_from_filename(path)
    if snapshot_date is None:
        return
    items, monthly_values = debt_bookings.parse_main_sheet(main_rows, main_sheet.title, path.name, "obvodny", snapshot_date)
    for item in items:
        red_columns = [6, 7, 9, 23, 24]
        green_columns = [8, 10]
        mark_row_columns(main_ws, item.source_row, red_columns, FILL_RED)
        mark_row_columns(main_ws, item.source_row, green_columns, FILL_GREEN)
    for item in monthly_values:
        mark_cell(main_ws, item.source_row, item.source_col, FILL_GREEN)
    deviation_sheet = debt_bookings.find_sheet(parsed_wb, "Отклонения")
    if deviation_sheet is not None and deviation_sheet.title in wb.sheetnames:
        ws = wb[deviation_sheet.title]
        for item in debt_bookings.parse_deviation_sheet(list(deviation_sheet.iter_rows(values_only=True)), deviation_sheet.title, path.name, "obvodny", snapshot_date):
            mark_row_columns(ws, item.source_row, [2, 3, 6, 10], FILL_RED)
            mark_row_columns(ws, item.source_row, [4, 5, 8, 9], FILL_GREEN)
    refusal_sheet = debt_bookings.find_sheet(parsed_wb, "Отказы")
    if refusal_sheet is not None and refusal_sheet.title in wb.sheetnames:
        ws = wb[refusal_sheet.title]
        for item in debt_bookings.parse_refusal_sheet(list(refusal_sheet.iter_rows(values_only=True)), refusal_sheet.title, path.name, "obvodny", snapshot_date):
            mark_row_columns(ws, item.source_row, [1, 4, 7, 9], FILL_RED)
            mark_row_columns(ws, item.source_row, [2, 3, 5, 6, 8], FILL_GREEN)


def mark_summary(path: Path, wb) -> None:
    _, _, _, cells = summary.parse_summary_file(path, SOURCE_ROOT / "сводная")
    for item in cells:
        if item.sheet_name not in wb.sheetnames:
            continue
        fill = FILL_RED if item.is_sensitive else FILL_BLUE
        mark_cell(wb[item.sheet_name], item.row_number, item.column_number, fill)


def apply_report_marks(path: Path, wb) -> str:
    text = str(path.relative_to(SOURCE_ROOT)).lower()
    if "платежный календарь" in text:
        mark_payment_calendar(path, wb)
        return "payment_calendar"
    if "дорожная карта" in text:
        mark_roadmap(path, wb)
        return "roadmap"
    if "3 модель" in text:
        mark_model(path, wb)
        return "model"
    if "не проектные расходы" in text:
        mark_non_project_expenses(path, wb)
        return "non_project_expenses"
    if "остатки в продаже" in text:
        mark_stock_for_sale(path, wb)
        return "stock_for_sale"
    if "отчет о продажах" in text:
        mark_sales_report(path, wb)
        return "sales_report"
    if "исполнении плана продаж" in text:
        mark_sales_plan_execution(path, wb)
        return "sales_plan_execution"
    if "отчет по агентам" in text:
        mark_agents_report(path, wb)
        return "agents_report"
    if "дз и брони" in text:
        mark_debt_bookings(path, wb)
        return "debt_and_bookings"
    if "сводная" in text:
        mark_summary(path, wb)
        return "summary"
    return "unknown"


def color_stats(wb) -> dict[str, int]:
    result = {"green": 0, "red": 0, "gray": 0, "blue": 0}
    colors = {
        "00C6EFCE": "green",
        "00FFC7CE": "red",
        "00D9D9D9": "gray",
        "00BDD7EE": "blue",
    }
    for ws in wb.worksheets:
        if ws.title == "Разметка":
            continue
        for row in ws.iter_rows():
            for cell in row:
                if not value_present(cell.value):
                    continue
                color = getattr(cell.fill.fgColor, "rgb", None)
                key = colors.get(color)
                if key:
                    result[key] += 1
    return result


def write_report(rows: list[dict[str, Any]]) -> None:
    path = OUTPUT_ROOT / "summary.md"
    lines = [
        "# Разметка оригинальных таблиц",
        "",
        "Цвета:",
        "- зеленый - участвует в пользовательской выдаче;",
        "- красный - есть в БД, но не выводится по безопасности;",
        "- серый - нет в БД и не выводится;",
        "- синий - есть в БД, но сейчас не используется в основном публичном ответе.",
        "",
        "| Отчет | Файл | Зеленый | Красный | Серый | Синий |",
        "| --- | --- | ---: | ---: | ---: | ---: |",
    ]
    for row in rows:
        lines.append(
            f"| {row['report']} | `{row['file']}` | {row['green']} | {row['red']} | {row['gray']} | {row['blue']} |",
        )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    if OUTPUT_ROOT.exists():
        shutil.rmtree(OUTPUT_ROOT)
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    report_rows = []
    files = sorted(path for path in SOURCE_ROOT.rglob("*.xlsx") if not path.name.startswith("~$"))
    for source_path in files:
        output_path = OUTPUT_ROOT / source_path.relative_to(SOURCE_ROOT)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb = load_workbook(source_path)
        mark_non_empty_gray(wb)
        report = apply_report_marks(source_path, wb)
        add_legend(wb)
        wb.save(output_path)
        stats = color_stats(wb)
        report_rows.append({"report": report, "file": str(source_path.relative_to(SOURCE_ROOT)), **stats})
    write_report(report_rows)
    print(f"created={len(report_rows)}")
    print(OUTPUT_ROOT)


if __name__ == "__main__":
    main()
